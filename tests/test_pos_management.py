from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import models
import schemas
import crud
from tests.conftest import TestingSessionLocal
from services import email as email_service


def _auth_headers(client: TestClient):
    resp = client.post(
        "/auth/login",
        json={"email": "master@kensar.com", "password": "2301"},
    )
    assert resp.status_code == 200
    token = resp.json()["token"]
    return {"Authorization": f"Bearer {token}"}


def test_settings_default_and_update(client: TestClient):
    headers = _auth_headers(client)
    response = client.get("/pos/settings", headers=headers)
    assert response.status_code == 200
    data = response.json()
    assert data["company_name"] == "Mi Negocio"
    assert data["theme_mode"] == "light"
    assert data["closure_email_recipients"] == []
    assert data["ticket_email_cc"] == []
    assert data["smtp_use_tls"] in (True, False)
    assert data["smtp_host"] == ""
    assert data["smtp_user"] == ""

    payload = {
        "company_name": "Kensar Labs",
        "tax_id": "123456789",
        "address": "Cra 1 #23-45",
        "contact_email": "info@example.com",
        "contact_phone": "3001234567",
        "theme_mode": "dark",
        "accent_color": "#123ABC",
        "ticket_footer": "Gracias por su compra",
        "auto_close_ticket": True,
        "low_stock_alert": False,
        "require_seller_pin": True,
        "notifications": {
            "daily_summary_email": True,
            "cash_alert_email": True,
            "cash_alert_sms": False,
            "monthly_report_email": True,
        },
        "logo_url": "https://example.com/logo.png",
        "ticket_logo_url": "/assets/logo.svg",
        "closure_email_recipients": ["cierres@kensar.com"],
        "ticket_email_cc": ["cc@kensar.com"],
        "smtp_host": "smtp.kensar.com",
        "smtp_port": 2525,
        "smtp_user": "smtp-user",
        "smtp_password": "smtp-pass",
        "smtp_use_tls": False,
        "email_from": "noreply@kensar.com",
    }

    update_response = client.put(
        "/pos/settings",
        json=payload,
        headers=headers,
    )
    assert update_response.status_code == 200
    updated = update_response.json()
    assert updated["company_name"] == "Kensar Labs"
    assert updated["accent_color"] == "#123ABC"
    assert updated["notifications"]["daily_summary_email"] is True
    assert updated["ticket_logo_url"] == "/assets/logo.svg"
    assert updated["closure_email_recipients"] == ["cierres@kensar.com"]
    assert updated["ticket_email_cc"] == ["cc@kensar.com"]
    assert updated["smtp_host"] == "smtp.kensar.com"
    assert updated["smtp_port"] == 2525
    assert updated["smtp_user"] == "smtp-user"
    assert updated["smtp_password"] == "smtp-pass"
    assert updated["smtp_use_tls"] is False
    assert updated["email_from"] == "noreply@kensar.com"

    refreshed = client.get("/pos/settings", headers=headers)
    assert refreshed.status_code == 200
    assert refreshed.json()["ticket_logo_url"] == "/assets/logo.svg"
    assert refreshed.json()["closure_email_recipients"] == ["cierres@kensar.com"]
    assert refreshed.json()["smtp_host"] == "smtp.kensar.com"


def test_settings_update_preserves_personalization_bindings_when_omitted(client: TestClient):
    headers = _auth_headers(client)
    baseline = client.get("/pos/settings", headers=headers).json()
    assert baseline.get("id")

    seeded_bindings = {
        "campana_clasica_mediana": {
            "product_id": "101",
            "product_sku": "CAMP-MED-BASE",
            "product_name": "Campana mediana",
            "product_slug": "campana-mediana",
            "service_id": "201",
            "service_sku": "SERV-CAMP-MED",
            "service_name": "Personalizacion campana mediana",
        },
        "campana_clasica_grande": {},
        "campana_cromada_mediana": {},
        "campana_cromada_grande": {},
        "guiro_mediano": {},
        "guiro_grande": {},
        "maraca_par": {},
    }
    seeded_payload = dict(baseline)
    seeded_payload["web_personalization_bindings"] = seeded_bindings
    seeded_payload.pop("id", None)
    seeded_response = client.put("/pos/settings", json=seeded_payload, headers=headers)
    assert seeded_response.status_code == 200
    assert (
        seeded_response.json()["web_personalization_bindings"]["campana_clasica_mediana"]["service_sku"]
        == "SERV-CAMP-MED"
    )

    # Simula payload parcial como el de dashboard/settings (sin web_personalization_bindings).
    partial_payload = {
        "company_name": "Kensar Labs Updated",
        "theme_mode": "light",
        "accent_color": "#0A84FF",
        "auto_close_ticket": False,
        "low_stock_alert": True,
        "require_seller_pin": False,
        "notifications": {
            "daily_summary_email": False,
            "cash_alert_email": False,
            "cash_alert_sms": False,
            "monthly_report_email": False,
        },
    }
    partial_response = client.put("/pos/settings", json=partial_payload, headers=headers)
    assert partial_response.status_code == 200
    persisted = client.get("/pos/settings", headers=headers)
    assert persisted.status_code == 200
    persisted_json = persisted.json()
    assert (
        persisted_json["web_personalization_bindings"]["campana_clasica_mediana"]["service_sku"]
        == "SERV-CAMP-MED"
    )


def test_create_and_update_pos_user(client: TestClient):
    create_payload = {
        "name": "Ana Pérez",
        "email": "ana@example.com",
        "role": "Administrador",
    }
    headers = _auth_headers(client)
    create_payload["password"] = "test1234"
    create_resp = client.post(
        "/pos/users",
        json=create_payload,
        headers=headers,
    )
    assert create_resp.status_code == 201
    user = create_resp.json()
    user_id = user["id"]
    assert user["status"] == "Activo"

    patch_resp = client.patch(
        f"/pos/users/{user_id}",
        json={"name": "Ana M. Pérez"},
        headers=headers,
    )
    assert patch_resp.status_code == 200
    assert patch_resp.json()["name"] == "Ana M. Pérez"

    fail_resp = client.patch(
        f"/pos/users/{user_id}",
        json={"status": "Inactivo"},
        headers=headers,
    )
    assert fail_resp.status_code == 400
    assert "Administrador" in fail_resp.json()["detail"]


def test_mobile_stock_binding_code_allows_pin_login_by_any_user_in_tenant(client: TestClient):
    headers = _auth_headers(client)

    user_a = client.post(
        "/pos/users",
        json={
            "name": "Operario Uno",
            "email": "operario1@example.com",
            "role": "Vendedor",
            "password": "test1234",
            "pin_plain": "4321",
        },
        headers=headers,
    )
    assert user_a.status_code == 201

    user_b = client.post(
        "/pos/users",
        json={
            "name": "Operario Dos",
            "email": "operario2@example.com",
            "role": "Supervisor",
            "password": "test5678",
            "pin_plain": "8765",
        },
        headers=headers,
    )
    assert user_b.status_code == 201

    setup_response = client.post(
        "/stock/devices/setup-code",
        json={"name": "Tablet Recepción Principal"},
        headers=headers,
    )
    assert setup_response.status_code == 201
    setup_payload = setup_response.json()
    setup_code = setup_payload["setup_code"]
    stock_device_id = setup_payload["device"]["id"]

    bind_response = client.post(
        "/auth/mobile-stock-bind",
        json={
            "setup_code": setup_code,
            "device_id": "STK-TABLET-01",
            "device_label": "Tablet recepción principal",
        },
    )
    assert bind_response.status_code == 200
    assert bind_response.json()["stock_device_id"] == stock_device_id

    reused_bind = client.post(
        "/auth/mobile-stock-bind",
        json={
            "setup_code": setup_code,
            "device_id": "STK-TABLET-01",
        },
    )
    assert reused_bind.status_code == 401

    login_response = client.post(
        "/auth/mobile-stock-login",
        json={
            "stock_device_id": stock_device_id,
            "pin": "8765",
            "device_id": "STK-TABLET-01",
            "device_label": "Tablet recepción principal",
        },
    )
    assert login_response.status_code == 200
    login_payload = login_response.json()
    assert login_payload["user"]["email"] == "operario2@example.com"
    assert login_payload["user"]["name"] == "Operario Dos"


def test_mobile_stock_legacy_email_flow_still_works(client: TestClient):
    headers = _auth_headers(client)

    create_user = client.post(
        "/pos/users",
        json={
            "name": "Operario Legado",
            "email": "operario.legacy@example.com",
            "role": "Vendedor",
            "password": "legacy1234",
            "pin_plain": "1357",
        },
        headers=headers,
    )
    assert create_user.status_code == 201

    login_response = client.post(
        "/auth/mobile-stock-login",
        json={
            "email": "operario.legacy@example.com",
            "pin": "1357",
            "device_id": "STK-LEGACY-01",
            "device_label": "Tablet legado",
        },
    )
    assert login_response.status_code == 200
    payload = login_response.json()
    assert payload["user"]["email"] == "operario.legacy@example.com"


def test_delete_inactive_stock_device_without_history(client: TestClient):
    headers = _auth_headers(client)

    setup_response = client.post(
        "/stock/devices/setup-code",
        json={"name": "Tablet temporal borrable"},
        headers=headers,
    )
    assert setup_response.status_code == 201
    stock_device_id = setup_response.json()["device"]["id"]

    deactivate_response = client.patch(
        f"/stock/devices/{stock_device_id}",
        json={"is_active": False},
        headers=headers,
    )
    assert deactivate_response.status_code == 200

    delete_response = client.delete(
        f"/stock/devices/{stock_device_id}",
        headers=headers,
    )
    assert delete_response.status_code == 204


def test_cannot_delete_active_stock_device(client: TestClient):
    headers = _auth_headers(client)

    setup_response = client.post(
        "/stock/devices/setup-code",
        json={"name": "Tablet activa no borrable"},
        headers=headers,
    )
    assert setup_response.status_code == 201
    stock_device_id = setup_response.json()["device"]["id"]

    delete_response = client.delete(
        f"/stock/devices/{stock_device_id}",
        headers=headers,
    )
    assert delete_response.status_code == 409
    assert "desactiva" in delete_response.json()["detail"].lower()


def test_receiving_documents_include_opening_and_closing_user_names(client: TestClient):
    headers = _auth_headers(client)

    creator_response = client.post(
        "/pos/users",
        json={
            "name": "Operario Apertura",
            "email": "operario.apertura@example.com",
            "role": "Vendedor",
            "password": "test1234",
            "pin_plain": "2468",
        },
        headers=headers,
    )
    assert creator_response.status_code == 201
    creator_id = creator_response.json()["id"]

    closer_response = client.post(
        "/pos/users",
        json={
            "name": "Operario Cierre",
            "email": "operario.cierre@example.com",
            "role": "Supervisor",
            "password": "test5678",
            "pin_plain": "9753",
        },
        headers=headers,
    )
    assert closer_response.status_code == 201
    closer_id = closer_response.json()["id"]

    db = TestingSessionLocal()
    try:
        tenant_id = crud.get_default_tenant_id(db)
        lot = models.ReceivingLot(
            tenant_id=tenant_id,
            lot_number="LOT-TRACE-001",
            status="closed",
            purchase_type="cash",
            origin_name="Tablet recepción principal",
            created_by_user_id=creator_id,
            closed_by_user_id=closer_id,
            created_at=datetime(2026, 7, 15, 8, 0, 0),
            closed_at=datetime(2026, 7, 15, 9, 0, 0),
        )
        db.add(lot)
        db.commit()
        db.refresh(lot)
    finally:
        db.close()

    response = client.get("/receiving/documents", headers=headers)
    assert response.status_code == 200
    items = response.json()["items"]
    payload = next(item for item in items if item["lot_number"] == "LOT-TRACE-001")
    assert payload["created_by_user_id"] == creator_id
    assert payload["created_by_user_name"] == "Operario Apertura"
    assert payload["closed_by_user_id"] == closer_id
    assert payload["closed_by_user_name"] == "Operario Cierre"


def _create_sale_record(
    customer_id=None,
    surcharge_amount: float = 0.0,
    surcharge_label: Optional[str] = None,
    pos_name: str = "POS 1",
):
    db = TestingSessionLocal()
    product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto cierre",
        price=100.0,
        cost=50.0,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add(product)
    db.commit()
    db.refresh(product)

    base_total = 100.0
    total_value = base_total + float(surcharge_amount or 0.0)
    sale_in = schemas.SaleCreate(
        payment_method="cash",
        total=total_value,
        paid_amount=total_value,
        change_amount=0.0,
        cart_discount_value=0.0,
        cart_discount_percent=0.0,
        surcharge_amount=surcharge_amount,
        surcharge_label=surcharge_label,
        customer_name="Test",
        notes=None,
        pos_name=pos_name,
        vendor_name="Tester",
        customer_id=customer_id,
        items=[
            schemas.SaleItemCreate(
                product_id=product.id,
                quantity=1,
                unit_price=100.0,
                product_sku=product.sku,
                product_name=product.name,
                product_barcode=product.barcode,
                discount=0.0,
            )
        ],
        payments=[schemas.SalePaymentCreate(method="cash", amount=total_value)],
    )
    sale = crud.create_sale(db, sale_in)
    sale_id = sale.id
    db.close()
    return sale_id


def _closure_payload():
    return {
        "pos_name": "POS 1",
        "counted_cash": 100.0,
        "notes": "Cierre automático",
    }


def _update_settings(client: TestClient, headers, **overrides):
    current = client.get("/pos/settings", headers=headers).json()
    current.pop("id", None)
    current.update(overrides)
    resp = client.put("/pos/settings", json=current, headers=headers)
    assert resp.status_code == 200
    return resp.json()


def test_closure_requires_pending_sales(client: TestClient):
    headers = _auth_headers(client)
    resp = client.post("/pos/closures", json=_closure_payload(), headers=headers)
    assert resp.status_code == 409
    assert "pendientes" in resp.json()["detail"]


def test_closure_marks_sales_and_prevents_duplicates(client: TestClient):
    headers = _auth_headers(client)
    sale_id = _create_sale_record()

    resp = client.post("/pos/closures", json=_closure_payload(), headers=headers)
    assert resp.status_code == 201
    data = resp.json()
    assert data["sales_count"] == 1
    assert data["total_surcharge"] == 0.0

    sale_resp = client.get(f"/pos/sales/{sale_id}", headers=headers)
    assert sale_resp.status_code == 200
    assert sale_resp.json()["closure_id"] == data["id"]

    resp_again = client.post("/pos/closures", json=_closure_payload(), headers=headers)
    assert resp_again.status_code == 409


def test_sale_with_surcharge_fields(client: TestClient):
    headers = _auth_headers(client)
    sale_id = _create_sale_record(surcharge_amount=15.5, surcharge_label="Addi")
    sale_resp = client.get(f"/pos/sales/{sale_id}", headers=headers)
    assert sale_resp.status_code == 200
    data = sale_resp.json()
    assert data["surcharge_amount"] == 15.5
    assert data["surcharge_label"] == "Addi"
    assert data["total"] == 115.5
    assert data["paid_amount"] == 115.5


def test_sale_total_is_recomputed_from_discounted_items(client: TestClient):
    headers = _auth_headers(client)
    db = TestingSessionLocal()
    product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto descuento total",
        price=100.0,
        cost=50.0,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add(product)
    db.commit()
    db.refresh(product)

    sale_in = schemas.SaleCreate(
        payment_method="cash",
        total=100.0,
        paid_amount=85.0,
        change_amount=0.0,
        cart_discount_value=5.0,
        cart_discount_percent=0.0,
        customer_name="Cliente prueba",
        notes=None,
        pos_name="POS 1",
        vendor_name="Tester",
        items=[
            schemas.SaleItemCreate(
                product_id=product.id,
                quantity=1,
                unit_price=100.0,
                product_sku=product.sku,
                product_name=product.name,
                product_barcode=product.barcode,
                discount=10.0,
            )
        ],
        payments=[schemas.SalePaymentCreate(method="cash", amount=85.0)],
    )
    sale = crud.create_sale(db, sale_in)
    db.close()

    assert sale.total == 85.0
    assert sale.paid_amount == 85.0
    assert sale.change_amount == 0.0
    assert sale.cart_discount_value == 5.0
    assert sale.items[0].line_discount_value == 10.0
    assert sale.items[0].total == 90.0


def test_closure_accumulates_surcharge(client: TestClient):
    headers = _auth_headers(client)
    isolated_pos_name = "POS SURCHARGE TEST"
    _create_sale_record(
        surcharge_amount=25.0,
        surcharge_label="Manual",
        pos_name=isolated_pos_name,
    )
    resp = client.post(
        "/pos/closures",
        json={
            "pos_name": isolated_pos_name,
            "counted_cash": 100.0,
            "notes": "Cierre automático",
        },
        headers=headers,
    )
    assert resp.status_code == 201
    data = resp.json()
    assert data["total_surcharge"] == 25.0


def test_closure_does_not_subtract_original_change_after_payment_adjustment(
    client: TestClient,
):
    headers = _auth_headers(client)
    isolated_pos_name = "POS AJUSTE PAGO NETO"
    db = TestingSessionLocal()
    try:
        product = models.Product(
            tenant_id=crud.get_default_tenant_id(db),
            name="Producto ajuste de pago cierre",
            price=32000.0,
            cost=16000.0,
            barcode=None,
            unit="UND",
            stock_min=0,
            preferred_qty=0,
            reorder_point=0,
            low_stock_alert=False,
            allow_price_change=False,
            active=True,
            service=False,
            includes_tax=False,
        )
        db.add(product)
        db.commit()
        db.refresh(product)

        sale = crud.create_sale(
            db,
            schemas.SaleCreate(
                payment_method="cash",
                total=32000.0,
                paid_amount=50000.0,
                change_amount=18000.0,
                cart_discount_value=0.0,
                cart_discount_percent=0.0,
                customer_name="Test ajuste",
                notes=None,
                pos_name=isolated_pos_name,
                vendor_name="Tester",
                items=[
                    schemas.SaleItemCreate(
                        product_id=product.id,
                        quantity=1,
                        unit_price=32000.0,
                        product_sku=product.sku,
                        product_name=product.name,
                        product_barcode=product.barcode,
                        discount=0.0,
                    )
                ],
                payments=[
                    schemas.SalePaymentCreate(method="cash", amount=50000.0)
                ],
            ),
        )
        sale_id = sale.id
    finally:
        db.close()

    adjustment_response = client.post(
        f"/pos/documents/sale/{sale_id}/adjust",
        json={
            "adjustment_type": "payment",
            "reason": "El pago neto se realizo por Nequi",
            "total_delta": 0.0,
            "payment_delta": -18000.0,
            "payload": {
                "payments": [{"method": "nequi", "amount": 32000.0}],
            },
        },
        headers=headers,
    )
    assert adjustment_response.status_code == 201

    preview_response = client.post(
        "/pos/closures/preview",
        json={
            "pos_name": isolated_pos_name,
            "counted_cash": 0.0,
            "notes": "Vista previa ajuste de pago",
        },
        headers=headers,
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["sales_count"] == 1
    assert preview["total_amount"] == 32000.0
    assert preview["net_amount"] == 32000.0
    assert preview["total_cash"] == 0.0
    assert preview["total_nequi"] == 32000.0
    assert preview["difference"] == 0.0
    assert sum(row["net"] for row in preview["methods_breakdown"]) == 32000.0

    closure_response = client.post(
        "/pos/closures",
        json={
            "pos_name": isolated_pos_name,
            "counted_cash": 0.0,
            "notes": "Cierre ajuste de pago",
        },
        headers=headers,
    )
    assert closure_response.status_code == 201
    closure_data = closure_response.json()
    for field in (
        "total_amount",
        "net_amount",
        "total_cash",
        "total_nequi",
        "total_refunds",
        "counted_cash",
        "difference",
        "methods_breakdown",
    ):
        assert closure_data[field] == preview[field]

def test_closure_separated_clarification_totals(client: TestClient):
    headers = _auth_headers(client)
    db = TestingSessionLocal()
    product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto separado cierre",
        sku="SEP-CLOSE-001",
        price=62000.0,
        cost=20000.0,
        barcode="SEP-CLOSE-001",
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add(product)
    db.commit()
    db.refresh(product)
    db.close()

    separated_payload = {
        "payment_method": "cash",
        "total": 25000.0,
        "paid_amount": 25000.0,
        "change_amount": 0.0,
        "cart_discount_value": 0.0,
        "cart_discount_percent": 0.0,
        "customer_name": "Cliente Separado Cierre",
        "notes": "Caso de prueba cierre separado",
        "pos_name": "POS 1",
        "vendor_name": "Tester",
        "items": [
            {
                "product_id": product.id,
                "quantity": 1,
                "unit_price": 62000.0,
                "product_sku": product.sku,
                "product_name": product.name,
                "product_barcode": product.barcode,
                "discount": 0.0,
            }
        ],
        "payments": [{"method": "cash", "amount": 10000.0}],
        "due_date": datetime.utcnow().isoformat(),
    }
    separated_resp = client.post("/separated-orders", json=separated_payload, headers=headers)
    assert separated_resp.status_code == 201
    separated_data = separated_resp.json()
    assert separated_data["total_amount"] == 62000.0
    assert separated_data["initial_payment"] == 25000.0
    assert separated_data["balance"] == 37000.0

    closure_resp = client.post("/pos/closures", json=_closure_payload(), headers=headers)
    assert closure_resp.status_code == 201
    closure_data = closure_resp.json()
    assert closure_data["net_amount"] == 25000.0
    assert closure_data["separated_summary"] is not None
    assert closure_data["separated_summary"]["tickets"] == 1
    assert closure_data["separated_summary"]["payments_total"] == 25000.0
    assert closure_data["separated_summary"]["reserved_total"] == 62000.0
    assert closure_data["separated_summary"]["pending_total"] == 37000.0
    assert closure_data["separated_summary"]["day_collected_total"] == 25000.0
    assert closure_data["separated_summary"]["day_with_pending_total"] == 62000.0
    assert closure_data["user_breakdown"]
    assert closure_data["user_breakdown"][0]["name"] == "Tester"
    assert closure_data["user_breakdown"][0]["total"] == 25000.0

    db = TestingSessionLocal()
    closure_row = (
        db.query(models.PosClosure)
        .filter(models.PosClosure.id == closure_data["id"])
        .first()
    )
    assert closure_row is not None
    closure_row.separated_summary = {
        "tickets": 1,
        "payments_total": 25000.0,
        "reserved_total": 62000.0,
        "pending_total": 40000.0,
        "day_collected_total": 25000.0,
        "day_with_pending_total": 65000.0,
    }
    db.commit()
    db.close()

    reloaded = client.get(f"/pos/closures/{closure_data['id']}", headers=headers)
    assert reloaded.status_code == 200
    reloaded_data = reloaded.json()
    assert reloaded_data["separated_summary"] is not None
    assert reloaded_data["separated_summary"]["pending_total"] == 37000.0
    assert reloaded_data["separated_summary"]["day_with_pending_total"] == 62000.0


def test_dashboard_summary_matches_collected_closure_total(client: TestClient):
    headers = _auth_headers(client)
    db = TestingSessionLocal()
    product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto dashboard cierre",
        sku="DASH-CLOSE-001",
        price=62000.0,
        cost=20000.0,
        barcode="DASH-CLOSE-001",
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add(product)
    db.commit()
    db.refresh(product)
    db.close()

    separated_payload = {
        "payment_method": "cash",
        "total": 25000.0,
        "paid_amount": 25000.0,
        "change_amount": 0.0,
        "cart_discount_value": 0.0,
        "cart_discount_percent": 0.0,
        "customer_name": "Cliente Dashboard",
        "notes": "Caso de prueba dashboard",
        "pos_name": "POS 1",
        "vendor_name": "Tester",
        "items": [
            {
                "product_id": product.id,
                "quantity": 1,
                "unit_price": 62000.0,
                "product_sku": product.sku,
                "product_name": product.name,
                "product_barcode": product.barcode,
                "discount": 0.0,
            }
        ],
        "payments": [{"method": "cash", "amount": 10000.0}],
        "due_date": datetime.utcnow().isoformat(),
    }
    separated_resp = client.post("/separated-orders", json=separated_payload, headers=headers)
    assert separated_resp.status_code == 201

    summary_resp = client.get("/dashboard/summary?source=metrik", headers=headers)
    assert summary_resp.status_code == 200
    summary = summary_resp.json()
    assert summary["today_sales_total"] == 25000.0
    assert summary["today_tickets"] == 1


def test_customer_crud_and_sales_association(client: TestClient):
    headers = _auth_headers(client)
    payload = {
        "name": "Cliente ACME",
        "phone": "3001234567",
        "email": "cliente@example.com",
        "tax_id": "900123456",
        "address": "Calle 123",
    }

    create_resp = client.post("/pos/customers", json=payload, headers=headers)
    assert create_resp.status_code == 201
    customer = create_resp.json()
    customer_id = customer["id"]

    list_resp = client.get("/pos/customers?search=ACME", headers=headers)
    assert list_resp.status_code == 200
    results = list_resp.json()
    assert any(entry["id"] == customer_id for entry in results)

    page_resp = client.get(
        "/pos/customers/page?search=ACME&status=active&segment=with_email&limit=8",
        headers=headers,
    )
    assert page_resp.status_code == 200
    customer_page = page_resp.json()
    assert customer_page["total"] == 1
    assert customer_page["summary"] == {
        "total": 1,
        "active": 1,
        "with_email": 1,
        "web_guests": 0,
    }
    assert [entry["id"] for entry in customer_page["items"]] == [customer_id]

    export_resp = client.get(
        "/pos/customers/export.xlsx?search=ACME&status=active&segment=with_email",
        headers=headers,
    )
    assert export_resp.status_code == 200
    assert export_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_resp.content), read_only=True)
    rows = list(workbook["Clientes"].iter_rows(values_only=True))
    assert rows[0][:4] == ("Nombre", "Teléfono", "Correo", "Documento")
    assert rows[1][:4] == (
        payload["name"],
        payload["phone"],
        payload["email"],
        payload["tax_id"],
    )
    assert len(rows) == 2

    update_resp = client.put(
        f"/pos/customers/{customer_id}",
        json={"phone": "3010000000"},
        headers=headers,
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["phone"] == "3010000000"

    sale_id = _create_sale_record(customer_id=customer_id)
    sale_resp = client.get(f"/pos/sales/{sale_id}", headers=headers)
    assert sale_resp.status_code == 200
    sale_data = sale_resp.json()
    assert sale_data["customer_id"] == customer_id
    assert sale_data["customer_name"] == customer["name"]
    assert sale_data["customer_phone"] == "3010000000"
    assert sale_data["customer_email"] == "cliente@example.com"
    assert sale_data["is_separated"] is False
    assert sale_data["initial_payment_method"] == sale_data["payment_method"]
    assert sale_data["initial_payment_amount"] == sale_data["paid_amount"]
    assert sale_data["balance"] is None
    assert sale_data["surcharge_amount"] == 0.0
    assert sale_data["surcharge_label"] is None

    delete_resp = client.delete(
        f"/pos/customers/{customer_id}",
        headers=headers,
    )
    assert delete_resp.status_code == 204

    list_after = client.get("/pos/customers?search=ACME", headers=headers)
    assert list_after.status_code == 200
    assert list_after.json() == []

    list_inactive = client.get(
        "/pos/customers?search=ACME&include_inactive=true",
        headers=headers,
    )
    assert list_inactive.status_code == 200
    assert len(list_inactive.json()) == 1
    assert list_inactive.json()[0]["is_active"] is False

    inactive_page_resp = client.get(
        "/pos/customers/page?search=ACME&status=inactive",
        headers=headers,
    )
    assert inactive_page_resp.status_code == 200
    inactive_page = inactive_page_resp.json()
    assert inactive_page["total"] == 1
    assert inactive_page["summary"]["active"] == 0


def test_send_sale_email_requires_recipients(client: TestClient, monkeypatch):
    headers = _auth_headers(client)
    _update_settings(
        client,
        headers,
        smtp_host="smtp.kensar.com",
        email_from="tickets@kensar.com",
        ticket_email_cc=["cc@example.com"],
    )
    sale_id = _create_sale_record()

    captured = {}

    def fake_send_email(**kwargs):
        captured.update(kwargs)

    monkeypatch.setattr(email_service, "send_email", fake_send_email)

    resp = client.post(
        f"/pos/sales/{sale_id}/email",
        json={"recipients": ["cliente@example.com"], "subject": "Ticket"},
        headers=headers,
    )
    assert resp.status_code == 200
    assert captured["recipients"] == ["cliente@example.com"]
    assert "Ticket" in captured["subject"]
    assert captured["smtp_config"]["smtp_host"] == "smtp.kensar.com"
    assert captured["cc"] == ["cc@example.com"]
    assert "Resumen de valores" in captured["html_body"]
    assert "Adjunto encontrarás el PDF" in captured["html_body"]
    assert "Efectivo" in captured["html_body"]
    assert "POS:" not in captured["html_body"]
    assert "Descuento carrito" not in captured["html_body"]
    assert "<strong>Productos</strong>" in captured["html_body"]
    assert "Nombre" in captured["html_body"]
    assert "Cantidad" in captured["html_body"]
    assert "Precio" in captured["html_body"]
    assert "Descuento" in captured["html_body"]
    assert "Total" in captured["html_body"]

    resp = client.post(
        f"/pos/sales/{sale_id}/email",
        json={},
        headers=headers,
    )
    assert resp.status_code == 400


def test_send_closure_email_uses_default_recipients(client: TestClient, monkeypatch):
    headers = _auth_headers(client)
    _update_settings(
        client,
        headers,
        closure_email_recipients=["cierres@example.com"],
        ticket_email_cc=[],
        smtp_host="smtp.kensar.com",
        email_from="cierres@kensar.com",
    )

    sale_id = _create_sale_record()
    resp = client.post("/pos/closures", json=_closure_payload(), headers=headers)
    assert resp.status_code == 201
    closure_id = resp.json()["id"]

    captured = {}

    def fake_send_email(**kwargs):
        captured.update(kwargs)

    monkeypatch.setattr(email_service, "send_email", fake_send_email)

    resp = client.post(
        f"/pos/closures/{closure_id}/email",
        json={},
        headers=headers,
    )
    assert resp.status_code == 200
    assert captured["recipients"] == ["cierres@example.com"]
    assert captured["smtp_config"]["email_from"] == "cierres@kensar.com"


def test_email_service_prefers_settings_config(monkeypatch):
    captured = {}

    class DummySMTP:
        def __init__(self, host, port):
            captured["host"] = host
            captured["port"] = port

        def starttls(self):
            captured["tls"] = True

        def login(self, user, password):
            captured["login"] = (user, password)

        def sendmail(self, sender, recipients, message):
            captured["sender"] = sender
            captured["recipients"] = recipients
            captured["message"] = message

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(email_service.smtplib, "SMTP", DummySMTP)

    config = {
        "smtp_host": "smtp.internal",
        "smtp_port": 2526,
        "smtp_user": "user",
        "smtp_password": "pass",
        "smtp_use_tls": False,
        "email_from": "noreply@example.com",
    }

    email_service.send_email(
        recipients=["cliente@example.com"],
        subject="Hola",
        html_body="<p>Test</p>",
        smtp_config=config,
    )

    assert captured["host"] == "smtp.internal"
    assert captured["port"] == 2526
    assert "tls" not in captured
    assert captured["login"] == ("user", "pass")
    assert captured["sender"] == "noreply@example.com"
    assert captured["recipients"] == ["cliente@example.com"]


def test_dashboard_monthly_sales_endpoint(client: TestClient):
    headers = _auth_headers(client)
    sale_id = _create_sale_record()

    db = TestingSessionLocal()
    sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    sale.created_at = datetime(2025, 1, 15)
    db.commit()
    db.close()

    resp = client.get("/dashboard/monthly-sales?year=2025", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) == 12
    january = next(item for item in data if item["month"] == 1)
    assert january["total"] > 0
    assert january["tickets"] == 1


def _create_change_test_sale(
    *,
    sale_name: str,
    sale_price: float,
    new_name: str,
    new_price: float,
):
    db = TestingSessionLocal()
    sale_product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name=sale_name,
        price=sale_price,
        cost=sale_price / 2,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    new_product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name=new_name,
        price=new_price,
        cost=new_price / 2,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add_all([sale_product, new_product])
    db.commit()
    db.refresh(sale_product)
    db.refresh(new_product)

    sale_in = schemas.SaleCreate(
        payment_method="cash",
        total=sale_price,
        paid_amount=sale_price,
        change_amount=0.0,
        cart_discount_value=0.0,
        cart_discount_percent=0.0,
        customer_name="Cliente prueba",
        notes=None,
        pos_name="POS 1",
        vendor_name="Tester",
        items=[
            schemas.SaleItemCreate(
                product_id=sale_product.id,
                quantity=1,
                unit_price=sale_price,
                product_sku=sale_product.sku,
                product_name=sale_product.name,
                product_barcode=sale_product.barcode,
                discount=0.0,
            )
        ],
        payments=[schemas.SalePaymentCreate(method="cash", amount=sale_price)],
    )
    sale = crud.create_sale(db, sale_in)
    sale_item_id = (
        db.query(models.SaleItem.id)
        .filter(models.SaleItem.sale_id == sale.id)
        .scalar()
    )
    result = (sale.id, sale_item_id, sale_product.id, new_product.id)
    db.close()
    return result


def _create_change_discounted_sale():
    db = TestingSessionLocal()
    sale_product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto cambio con descuento",
        price=10000.0,
        cost=5000.0,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    other_product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Producto extra cambio con descuento",
        price=20000.0,
        cost=10000.0,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    new_product = models.Product(
        tenant_id=crud.get_default_tenant_id(db),
        name="Nuevo producto cambio con descuento",
        price=11000.0,
        cost=5500.0,
        barcode=None,
        unit="UND",
        stock_min=0,
        preferred_qty=0,
        reorder_point=0,
        low_stock_alert=False,
        allow_price_change=False,
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add_all([sale_product, other_product, new_product])
    db.commit()
    db.refresh(sale_product)
    db.refresh(other_product)
    db.refresh(new_product)

    sale_in = schemas.SaleCreate(
        payment_method="cash",
        total=29999.0,
        paid_amount=29999.0,
        change_amount=0.0,
        cart_discount_value=1.0,
        cart_discount_percent=0.0,
        customer_name="Cliente prueba",
        notes=None,
        pos_name="POS 1",
        vendor_name="Tester",
        items=[
            schemas.SaleItemCreate(
                product_id=sale_product.id,
                quantity=1,
                unit_price=sale_product.price,
                product_sku=sale_product.sku,
                product_name=sale_product.name,
                product_barcode=sale_product.barcode,
                discount=0.0,
            ),
            schemas.SaleItemCreate(
                product_id=other_product.id,
                quantity=1,
                unit_price=other_product.price,
                product_sku=other_product.sku,
                product_name=other_product.name,
                product_barcode=other_product.barcode,
                discount=0.0,
            ),
        ],
        payments=[schemas.SalePaymentCreate(method="cash", amount=29999.0)],
    )
    sale = crud.create_sale(db, sale_in)
    sale_item_id = (
        db.query(models.SaleItem.id)
        .filter(models.SaleItem.sale_id == sale.id)
        .filter(models.SaleItem.product_id == sale_product.id)
        .scalar()
    )
    result = (sale.id, sale_item_id, sale_product.id, new_product.id)
    db.close()
    return result


def test_return_creates_inventory_entry(client: TestClient):
    headers = _auth_headers(client)
    sale_id, sale_item_id, sale_product_id, _ = _create_change_test_sale(
        sale_name="Producto devolución",
        sale_price=15000.0,
        new_name="Nuevo no usado",
        new_price=15000.0,
    )

    resp = client.post(
        "/pos/returns",
        json={
            "sale_id": sale_id,
            "items": [
                {"sale_item_id": sale_item_id, "quantity": 1, "reason": "cambio"},
            ],
        },
        headers=headers,
    )
    assert resp.status_code == 201
    return_id = resp.json()["id"]

    db = TestingSessionLocal()
    movements = (
        db.query(models.InventoryMovement)
        .filter(models.InventoryMovement.reference_type == "sale_return")
        .filter(models.InventoryMovement.reference_id == return_id)
        .all()
    )
    db.close()

    assert len(movements) == 1
    assert movements[0].product_id == sale_product_id
    assert movements[0].qty_delta == 1
    assert movements[0].reason == "transfer_in"


def test_change_creates_and_voids_inventory_movements(client: TestClient):
    headers = _auth_headers(client)
    sale_id, sale_item_id, sale_product_id, new_product_id = _create_change_test_sale(
        sale_name="Producto cambio",
        sale_price=20000.0,
        new_name="Producto nuevo cambio",
        new_price=20000.0,
    )

    resp = client.post(
        "/pos/changes",
        json={
            "sale_id": sale_id,
            "return_items": [
                {"sale_item_id": sale_item_id, "quantity": 1, "reason": "cambio"},
            ],
            "new_items": [
                {"product_id": new_product_id, "quantity": 1},
            ],
        },
        headers=headers,
    )
    assert resp.status_code == 201
    change_id = resp.json()["id"]

    db = TestingSessionLocal()
    movements = (
        db.query(models.InventoryMovement)
        .filter(models.InventoryMovement.reference_type == "sale_change")
        .filter(models.InventoryMovement.reference_id == change_id)
        .order_by(models.InventoryMovement.id.asc())
        .all()
    )
    assert len(movements) == 2
    assert {movement.product_id for movement in movements} == {sale_product_id, new_product_id}
    assert any(movement.product_id == sale_product_id and movement.qty_delta == 1 for movement in movements)
    assert any(
        movement.product_id == new_product_id and movement.qty_delta == -1
        for movement in movements
    )

    void_resp = client.post(
        f"/pos/changes/{change_id}/void",
        json={"reason": "Corrección"},
        headers=headers,
    )
    assert void_resp.status_code == 200

    all_movements = (
        db.query(models.InventoryMovement)
        .filter(models.InventoryMovement.reference_type == "sale_change")
        .filter(models.InventoryMovement.reference_id == change_id)
        .order_by(models.InventoryMovement.id.asc())
        .all()
    )
    db.close()

    assert len(all_movements) == 4
    assert sum(m.qty_delta for m in all_movements if m.product_id == sale_product_id) == 0
    assert sum(m.qty_delta for m in all_movements if m.product_id == new_product_id) == 0


def test_change_accepts_discounted_ticket_with_integer_payment(client: TestClient):
    headers = _auth_headers(client)
    sale_id, sale_item_id, _, new_product_id = _create_change_discounted_sale()

    resp = client.post(
        "/pos/changes",
        json={
            "sale_id": sale_id,
            "return_items": [
                {"sale_item_id": sale_item_id, "quantity": 1, "reason": "cambio"},
            ],
            "new_items": [
                {"product_id": new_product_id, "quantity": 1},
            ],
            "payments": [
                {"method": "cash", "amount": 1000},
            ],
        },
        headers=headers,
    )
    assert resp.status_code == 201
    data = resp.json()
    assert data["total_credit"] == 10000
    assert data["total_new"] == 11000
    assert data["extra_payment"] == 1000
    assert data["refund_due"] == 0
    assert data["payments"][0]["amount"] == 1000


def test_chained_change_can_be_returned_once_and_preserves_lineage(client: TestClient):
    headers = _auth_headers(client)
    sale_id, sale_item_id, _, first_new_product_id = _create_change_test_sale(
        sale_name="Producto cadena original",
        sale_price=20000.0,
        new_name="Producto cadena cambio uno",
        new_price=20000.0,
    )
    first = client.post(
        "/pos/changes",
        json={
            "sale_id": sale_id,
            "return_items": [{"sale_item_id": sale_item_id, "quantity": 1}],
            "new_items": [{"product_id": first_new_product_id, "quantity": 1}],
        },
        headers=headers,
    )
    assert first.status_code == 201, first.text
    first_data = first.json()
    first_source_id = first_data["items_new"][0]["id"]

    db = TestingSessionLocal()
    tenant_id = crud.get_default_tenant_id(db)
    second_product = models.Product(
        tenant_id=tenant_id,
        name="Producto cadena cambio dos",
        price=18000.0,
        cost=9000.0,
        unit="UND",
        active=True,
        service=False,
        includes_tax=False,
    )
    db.add(second_product)
    db.commit()
    second_product_id = second_product.id
    db.close()

    second = client.post(
        "/pos/changes",
        json={
            "sale_id": sale_id,
            "source_document_number": first_data["document_number"],
            "return_items": [{
                "source_type": "change",
                "source_item_id": first_source_id,
                "quantity": 1,
            }],
            "new_items": [{"product_id": second_product_id, "quantity": 1}],
            "refund_method": "card",
        },
        headers=headers,
    )
    assert second.status_code == 201, second.text
    second_data = second.json()
    assert second_data["source_document_type"] == "change"
    assert second_data["source_document_number"] == first_data["document_number"]
    assert second_data["refund_due"] == 2000
    assert second_data["refund_method"] == "card"

    duplicate = client.post(
        "/pos/returns",
        json={
            "sale_id": sale_id,
            "items": [{
                "source_type": "change",
                "source_item_id": first_source_id,
                "quantity": 1,
            }],
            "payments": [{"method": "cash", "amount": 20000}],
        },
        headers=headers,
    )
    assert duplicate.status_code == 400
    assert "cantidad disponible" in duplicate.json()["detail"].lower()

    resolved_first = client.get(
        "/pos/operation-documents/resolve",
        params={"code": first_data["document_number"]},
        headers=headers,
    )
    assert resolved_first.status_code == 200
    assert resolved_first.json()["items"][0]["available_quantity"] == 0
    assert resolved_first.json()["allowed_actions"] == []

    second_source_id = second_data["items_new"][0]["id"]
    returned = client.post(
        "/pos/returns",
        json={
            "sale_id": sale_id,
            "source_document_number": second_data["document_number"],
            "items": [{
                "source_type": "change",
                "source_item_id": second_source_id,
                "quantity": 1,
            }],
            "payments": [{"method": "card", "amount": 18000}],
        },
        headers=headers,
    )
    assert returned.status_code == 201, returned.text
    returned_data = returned.json()
    assert returned_data["source_document_type"] == "change"
    assert returned_data["source_document_number"] == second_data["document_number"]
    assert returned_data["total_refund"] == 18000

    resolved_chain = client.get(
        "/pos/operation-documents/resolve",
        params={"code": returned_data["document_number"]},
        headers=headers,
    )
    assert resolved_chain.status_code == 200
    chain = resolved_chain.json()["chain"]
    assert [entry["document_type"] for entry in chain] == [
        "sale", "change", "change", "return"
    ]
    assert [(item["action"], item["product_name"], item["quantity"]) for item in chain[0]["items"]] == [
        ("purchased", "Producto cadena original", 1)
    ]
    assert [(item["action"], item["product_name"]) for item in chain[1]["items"]] == [
        ("returned", "Producto cadena original"),
        ("received", "Producto cadena cambio uno"),
    ]
    assert [(item["action"], item["product_name"]) for item in chain[2]["items"]] == [
        ("returned", "Producto cadena cambio uno"),
        ("received", "Producto cadena cambio dos"),
    ]
    assert [(item["action"], item["product_name"]) for item in chain[3]["items"]] == [
        ("returned", "Producto cadena cambio dos")
    ]

    cannot_void_parent = client.post(
        f"/pos/changes/{first_data['id']}/void",
        json={"reason": "No debe romper la cadena"},
        headers=headers,
    )
    assert cannot_void_parent.status_code == 400
    assert "ya fue cambiado o devuelto" in cannot_void_parent.json()["detail"]
