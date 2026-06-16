from datetime import datetime
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

import crud
import models
import schemas
from database import get_db
from dependencies import require_module_access, require_permission


router = APIRouter(
    prefix="/comercio-web",
    tags=["comercio-web"],
)


def _tenant_id_for_user(db: Session, user: models.PosUser) -> int:
    tenant_id = crud.resolve_user_tenant_id(db, user)
    if tenant_id is None:
        raise HTTPException(status_code=403, detail="Usuario sin empresa asignada")
    return tenant_id


@router.get("/orders", response_model=list[schemas.WebOrderRead])
def list_comercio_web_orders(
    limit: int = Query(default=100, ge=1, le=200),
    status: Optional[schemas.WebOrderStatus] = Query(default=None),
    payment_status: Optional[schemas.WebOrderPaymentStatus] = Query(default=None),
    search: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    orders = crud.list_backoffice_web_orders(
        db,
        tenant_id=tenant_id,
        limit=limit,
        status=status,
        payment_status=payment_status,
        search=search,
    )
    return [crud._serialize_web_order(order) for order in orders]


@router.get("/catalog/products", response_model=list[schemas.ProductRead])
def list_comercio_web_catalog_products(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=1000, ge=1, le=5000),
    q: Optional[str] = Query(default=None),
    published_only: Optional[bool] = Query(default=None),
    configured_only: Optional[bool] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    if (
        q is not None
        or published_only is not None
        or configured_only is not None
        or skip != 0
        or limit != 1000
    ):
        return crud.search_comercio_web_catalog_products(
            db,
            tenant_id=tenant_id,
            q=q,
            published_only=published_only,
            configured_only=configured_only,
            skip=skip,
            limit=limit,
        )
    return crud.get_products(db, skip=skip, limit=limit, tenant_id=tenant_id)


@router.get("/catalog/technical-spec-types", response_model=list[str])
def list_comercio_web_technical_spec_types(
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    _tenant_id_for_user(db, current_user)
    return crud.get_comercio_web_technical_spec_types()


@router.get(
    "/catalog/publications",
    response_model=schemas.ComercioWebCatalogPublicationPage,
)
def list_comercio_web_publications(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    q: Optional[str] = Query(default=None),
    field: str = Query(default="all"),
    status_filter: str = Query(default="all"),
    featured_filter: str = Query(default="all"),
    badge_filter: str = Query(default="all"),
    stock_filter: str = Query(default="all"),
    category_key: Optional[str] = Query(default=None),
    subcategory_key: Optional[str] = Query(default=None),
    order: str = Query(default="newest"),
    active_only: bool = Query(default=True),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_publications_page(
        db,
        tenant_id=tenant_id,
        q=q,
        field=field,
        status_filter=status_filter,
        featured_filter=featured_filter,
        badge_filter=badge_filter,
        stock_filter=stock_filter,
        category_key=category_key,
        subcategory_key=subcategory_key,
        order=order,
        active_only=active_only,
        skip=skip,
        limit=limit,
    )


@router.get("/catalog/publications/export/xlsx")
def export_comercio_web_publications_xlsx(
    q: Optional[str] = Query(default=None),
    field: str = Query(default="all"),
    status_filter: str = Query(default="all"),
    featured_filter: str = Query(default="all"),
    badge_filter: str = Query(default="all"),
    stock_filter: str = Query(default="all"),
    category_key: Optional[str] = Query(default=None),
    subcategory_key: Optional[str] = Query(default=None),
    order: str = Query(default="newest"),
    active_only: bool = Query(default=True),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    page = crud.list_comercio_web_publications_page(
        db,
        tenant_id=tenant_id,
        q=q,
        field=field,
        status_filter=status_filter,
        featured_filter=featured_filter,
        badge_filter=badge_filter,
        stock_filter=stock_filter,
        category_key=category_key,
        subcategory_key=subcategory_key,
        order=order,
        active_only=active_only,
        skip=0,
        limit=200000,
    )

    categories = crud.list_comercio_web_catalog_categories(
        db,
        tenant_id=tenant_id,
        include_inactive=True,
    )
    category_name_by_key = {
        (row.key or "").strip().lower(): row.name for row in categories if row.key
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Publicaciones"
    cop_number_format = '"$" #,##0'
    headers = [
        "ID",
        "Nombre web",
        "Nombre base",
        "Slug web",
        "SKU",
        "Código barras",
        "Marca",
        "Grupo",
        "Proveedor",
        "Categoría web key",
        "Categoría web nombre",
        "Precio base",
        "Costo",
        "Precio web calculado",
        "Precio comparar",
        "Fuente precio web",
        "Valor precio web",
        "Modo precio web",
        "Estado web",
        "Publicado web",
        "Activo inventario",
        "Destacado",
        "Badge",
        "Orden web",
        "Visible sin stock",
        "Servicio",
        "Unidad",
        "Imagen principal",
        "Imagen miniatura",
        "Galería",
        "Descripción corta",
        "Descripción larga",
        "Mensaje WhatsApp",
        "Garantía",
        "Publicado en",
        "Actualizado en",
    ]
    sheet.append(headers)
    money_columns = {
        "Precio base": None,
        "Costo": None,
        "Precio web calculado": None,
        "Precio comparar": None,
        "Valor precio web": None,
    }
    for index, header in enumerate(headers, start=1):
        if header in money_columns:
            money_columns[header] = index

    for product in page.get("items", []):
        category_key_value = (product.web_category_key or "").strip()
        category_name = category_name_by_key.get(category_key_value.lower(), category_key_value)
        gallery_value = product.web_gallery_urls
        if isinstance(gallery_value, list):
            gallery_text = " | ".join(str(item).strip() for item in gallery_value if str(item).strip())
        else:
            gallery_text = str(gallery_value or "").strip()
        sale_price = float(crud.resolve_web_product_sale_price(product))
        compare_price = (
            float(product.web_compare_price)
            if product.web_compare_price is not None
            else None
        )
        row_values = [
            int(product.id),
            (product.web_name or "").strip(),
            (product.name or "").strip(),
            (product.web_slug or "").strip(),
            (product.sku or "").strip(),
            (product.barcode or "").strip(),
            (product.brand or "").strip(),
            (product.group_name or "").strip(),
            (product.supplier or "").strip(),
            category_key_value,
            category_name or "",
            float(product.price or 0),
            float(product.cost or 0),
            sale_price,
            compare_price,
            (product.web_price_source or "base").strip(),
            float(product.web_price_value) if product.web_price_value is not None else None,
            (product.web_price_mode or "visible").strip(),
            "publicado" if bool(product.web_published) else "pausado",
            "si" if bool(product.web_published) else "no",
            "si" if bool(product.active) else "no",
            "si" if bool(product.web_featured) else "no",
            (product.web_badge_text or "").strip(),
            int(product.web_sort_order or 0),
            "si" if bool(product.web_visible_when_out_of_stock) else "no",
            "si" if bool(product.service) else "no",
            (product.unit or "").strip(),
            (product.image_url or "").strip(),
            (product.image_thumb_url or "").strip(),
            gallery_text,
            (product.web_short_description or "").strip(),
            (product.web_long_description or "").strip(),
            (product.web_whatsapp_message or "").strip(),
            (product.web_warranty_text or "").strip(),
            product.web_published_at.isoformat() if product.web_published_at else "",
            product.updated_at.isoformat() if product.updated_at else "",
        ]
        sheet.append(row_values)
        current_row = sheet.max_row
        for col_index in money_columns.values():
            if not col_index:
                continue
            sheet.cell(row=current_row, column=col_index).number_format = cop_number_format

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"catalogo_web_publicaciones_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/catalog/combos", response_model=list[schemas.ComercioWebComboRead])
def list_comercio_web_combos(
    q: Optional[str] = Query(default=None),
    published_only: Optional[bool] = Query(default=None),
    active_only: Optional[bool] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_combos(
        db,
        tenant_id=tenant_id,
        q=q,
        published_only=published_only,
        active_only=active_only,
    )


@router.post("/catalog/combos", response_model=schemas.ComercioWebComboRead, status_code=201)
def create_comercio_web_combo(
    payload: schemas.ComercioWebComboCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.create_comercio_web_combo(db, tenant_id=tenant_id, payload=payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put("/catalog/combos/{combo_id}", response_model=schemas.ComercioWebComboRead)
def update_comercio_web_combo(
    combo_id: int,
    payload: schemas.ComercioWebComboUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.update_comercio_web_combo(
            db,
            tenant_id=tenant_id,
            combo_id=combo_id,
            payload=payload,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.delete("/catalog/combos/{combo_id}", status_code=204)
def delete_comercio_web_combo(
    combo_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        crud.delete_comercio_web_combo(db, tenant_id=tenant_id, combo_id=combo_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return None


@router.put("/catalog/products/{product_id}", response_model=schemas.ProductRead)
def update_comercio_web_catalog_product(
    product_id: int,
    payload: schemas.ProductUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    product = crud.get_product(db, product_id, tenant_id=tenant_id)
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    try:
        updated = crud.update_product(db, product, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    crud.create_product_audit_log(
        db,
        product_id=updated.id,
        action="update",
        actor_user=current_user,
        changes={"commerce_web_catalog": True},
    )
    return updated


@router.get(
    "/catalog/categories",
    response_model=list[schemas.ComercioWebCatalogCategoryRead],
)
def list_comercio_web_catalog_categories(
    include_inactive: bool = Query(default=True),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_catalog_categories(
        db,
        tenant_id=tenant_id,
        include_inactive=include_inactive,
    )


@router.post(
    "/catalog/categories",
    response_model=schemas.ComercioWebCatalogCategoryRead,
)
def create_comercio_web_catalog_category(
    payload: schemas.ComercioWebCatalogCategoryCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.create_comercio_web_catalog_category(
            db,
            tenant_id=tenant_id,
            payload=payload,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put(
    "/catalog/categories/{category_id}",
    response_model=schemas.ComercioWebCatalogCategoryRead,
)
def update_comercio_web_catalog_category(
    category_id: int,
    payload: schemas.ComercioWebCatalogCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.update_comercio_web_catalog_category(
            db,
            tenant_id=tenant_id,
            category_id=category_id,
            payload=payload,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrada" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc


@router.delete("/catalog/categories/{category_id}")
def delete_comercio_web_catalog_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        crud.delete_comercio_web_catalog_category(
            db,
            tenant_id=tenant_id,
            category_id=category_id,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrada" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc
    return {"ok": True}


@router.get(
    "/home-sliders",
    response_model=list[schemas.ComercioWebHomeSliderRead],
)
def list_comercio_web_home_sliders(
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_home_sliders(db, tenant_id=tenant_id)


@router.put(
    "/home-sliders/{slot}",
    response_model=schemas.ComercioWebHomeSliderRead,
)
def update_comercio_web_home_slider(
    slot: int,
    payload: schemas.ComercioWebHomeSliderUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.update_comercio_web_home_slider(
            db,
            tenant_id=tenant_id,
            slot=slot,
            payload=payload,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrado" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc


@router.get(
    "/catalog/discount-codes/{discount_code_id}/usage",
    response_model=schemas.ComercioWebDiscountCodeUsagePage,
)
def list_comercio_web_discount_code_usage(
    discount_code_id: int,
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.list_comercio_web_discount_code_usage_page(
            db,
            tenant_id=tenant_id,
            discount_code_id=discount_code_id,
            skip=skip,
            limit=limit,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.get(
    "/catalog/description-templates",
    response_model=list[schemas.ComercioWebDescriptionTemplateRead],
)
def list_comercio_web_description_templates(
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_description_templates(
        db,
        tenant_id=tenant_id,
    )


@router.post(
    "/catalog/description-templates",
    response_model=schemas.ComercioWebDescriptionTemplateRead,
)
def create_comercio_web_description_template(
    payload: schemas.ComercioWebDescriptionTemplateCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.create_comercio_web_description_template(
            db,
            tenant_id=tenant_id,
            payload=payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put(
    "/catalog/description-templates/{template_key}",
    response_model=schemas.ComercioWebDescriptionTemplateRead,
)
def update_comercio_web_description_template(
    template_key: str,
    payload: schemas.ComercioWebDescriptionTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.update_comercio_web_description_template(
            db,
            tenant_id=tenant_id,
            template_key=template_key,
            payload=payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrada" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc


@router.delete("/catalog/description-templates/{template_key}")
def delete_comercio_web_description_template(
    template_key: str,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        crud.delete_comercio_web_description_template(
            db,
            tenant_id=tenant_id,
            template_key=template_key,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrada" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc
    return {"ok": True}


@router.post(
    "/catalog/description-templates/reset",
    response_model=list[schemas.ComercioWebDescriptionTemplateRead],
)
def reset_comercio_web_description_templates(
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.reset_comercio_web_description_templates(
        db,
        tenant_id=tenant_id,
        actor_user_id=current_user.id,
    )


@router.get(
    "/catalog/discount-codes",
    response_model=schemas.ComercioWebDiscountCodePage,
)
def list_comercio_web_discount_codes(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    q: Optional[str] = Query(default=None),
    active_only: Optional[bool] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    return crud.list_comercio_web_discount_codes_page(
        db,
        tenant_id=tenant_id,
        q=q,
        active_only=active_only,
        skip=skip,
        limit=limit,
    )


@router.post(
    "/catalog/discount-codes",
    response_model=schemas.ComercioWebDiscountCodeRead,
)
def create_comercio_web_discount_code(
    payload: schemas.ComercioWebDiscountCodeCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.create_comercio_web_discount_code(
            db,
            tenant_id=tenant_id,
            payload=payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put(
    "/catalog/discount-codes/{discount_code_id}",
    response_model=schemas.ComercioWebDiscountCodeRead,
)
def update_comercio_web_discount_code(
    discount_code_id: int,
    payload: schemas.ComercioWebDiscountCodeUpdate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    try:
        return crud.update_comercio_web_discount_code(
            db,
            tenant_id=tenant_id,
            discount_code_id=discount_code_id,
            payload=payload,
        )
    except ValueError as exc:
        detail = str(exc)
        if "no encontrado" in detail.lower():
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc


@router.get("/orders/{order_id}", response_model=schemas.WebOrderRead)
def get_comercio_web_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.view")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    order = crud.get_backoffice_web_order(db, order_id, tenant_id=tenant_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden web no encontrada")
    order = refresh_backoffice_order_payment_statuses(db, [order])[0]
    return crud._serialize_web_order(order)


@router.post("/orders/{order_id}/payments", response_model=schemas.WebOrderRead)
def record_comercio_web_payment(
    order_id: int,
    payload: schemas.WebOrderPaymentRecordRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    order = crud.get_backoffice_web_order(db, order_id, tenant_id=tenant_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden web no encontrada")
    try:
        return crud.record_web_order_payment(
            db,
            order,
            payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/orders/{order_id}/status", response_model=schemas.WebOrderRead)
def update_comercio_web_order_status(
    order_id: int,
    payload: schemas.WebOrderStatusUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    order = crud.get_backoffice_web_order(db, order_id, tenant_id=tenant_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden web no encontrada")
    try:
        return crud.update_backoffice_web_order_status(
            db,
            order,
            payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/orders/{order_id}/convert-to-sale", response_model=schemas.WebOrderRead)
def convert_comercio_web_order_to_sale(
    order_id: int,
    payload: schemas.WebOrderConvertToSaleRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("commerce_web.manage")),
    _: models.PosUser = Depends(require_module_access("commerce_web")),
):
    tenant_id = _tenant_id_for_user(db, current_user)
    order = crud.get_backoffice_web_order(db, order_id, tenant_id=tenant_id)
    if not order:
        raise HTTPException(status_code=404, detail="Orden web no encontrada")
    try:
        return crud.convert_web_order_to_sale(
            db,
            order,
            payload,
            actor_user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
