from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from io import BytesIO
from typing import Optional
import json
import urllib.error
import urllib.request
import ssl
import os
import threading
import logging
import uuid

try:
    import certifi
except ModuleNotFoundError:
    certifi = None

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

import schemas
from dependencies import require_permission


router = APIRouter(
    prefix="/labels",
    tags=["labels"],
)
DEFAULT_SATO_CLOUD_TIMEOUT_SECONDS = int(
    os.getenv("SATO_CLOUD_TIMEOUT_SECONDS", "20")
)
logger = logging.getLogger("kensar.labels")


def _load_sato_cloud_settings() -> tuple[str, str, bool]:
    base_url = os.getenv(
        "SATO_CLOUD_BASE_URL",
        "https://satoeasyprint.com/api/v1/labels/print",
    ).strip().rstrip("/")
    client_key = os.getenv("SATO_CLOUD_CLIENT_KEY", "").strip()
    skip_tls_verify = (
        os.getenv("SATO_CLOUD_SKIP_TLS_VERIFY", "false").strip().lower()
        in {"1", "true", "yes", "on"}
    )
    return base_url, client_key, skip_tls_verify


def _format_price_text(value: float) -> str:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise ValueError("Precio inválido")

    quantized = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    integer_part, fractional_part = f"{quantized:.2f}".split(".")
    integer_formatted = f"{int(integer_part):,}".replace(",", ".")
    if fractional_part == "00":
        return f"${integer_formatted}"
    return f"${integer_formatted},{fractional_part}"


def _resolve_sku_text(sku: Optional[str], product_id: int) -> str:
    value = sku.strip() if sku else ""
    code = value or str(product_id)
    return f"Code: {code}"


@router.post("/export/xlsx")
def export_labels_excel(
    payload: schemas.LabelExportRequest,
    _: object = Depends(require_permission("labels.export")),
):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Debes enviar al menos un producto")

    workbook = Workbook()
    try:
        workbook.calculation_properties.fullCalcOnLoad = True
    except AttributeError:
        pass
    sheet = workbook.active
    sheet.title = "Etiquetas"
    headers = ["SKU", "Nombre", "Precio", "Código de barras"]
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.number_format = "@"
        cell.data_type = "s"
        if cell.value is not None:
            cell.value = str(cell.value)

    for item in payload.items:
        try:
            price_text = _format_price_text(item.price)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        repetitions = max(item.quantity, 1)
        for _ in range(repetitions):
            sheet.append(
                [
                    _resolve_sku_text(item.sku, item.product_id),
                    item.name,
                    price_text,
                    (item.barcode or ""),
                ]
            )
            for cell in sheet[sheet.max_row]:
                cell.number_format = "@"
                cell.data_type = "s"
                if cell.value is not None:
                    cell.value = str(cell.value)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=labels.xlsx",
        },
    )


@router.post("/cloud/print/{serial}")
def print_label_via_sato_cloud(
    serial: str,
    payload: schemas.LabelCloudPrintRequest,
    _: object = Depends(require_permission("labels.export")),
):
    request_id = uuid.uuid4().hex[:12]
    serial_value = serial.strip()
    if not serial_value:
        raise HTTPException(status_code=400, detail="Serial inválido")

    sato_cloud_base_url, sato_cloud_client_key, sato_cloud_skip_tls_verify = (
        _load_sato_cloud_settings()
    )

    if not sato_cloud_base_url:
        raise HTTPException(
            status_code=500,
            detail="Falta configurar SATO_CLOUD_BASE_URL en el backend.",
        )
    if not sato_cloud_client_key:
        raise HTTPException(
            status_code=500,
            detail="Falta configurar SATO_CLOUD_CLIENT_KEY en el backend.",
        )

    target_url = f"{sato_cloud_base_url}/{serial_value}"
    request_body = json.dumps([payload.payload.model_dump()]).encode("utf-8")
    request = urllib.request.Request(
        target_url,
        data=request_body,
        headers={
            "Content-Type": "application/json",
            "X-CLIENT-KEY": sato_cloud_client_key,
        },
        method="POST",
    )

    try:
        if sato_cloud_skip_tls_verify:
            tls_context = ssl._create_unverified_context()
        else:
            if certifi is not None:
                tls_context = ssl.create_default_context(cafile=certifi.where())
            else:
                tls_context = ssl.create_default_context()
    except Exception as exc:
        logger.exception(
            "labels.cloud[%s] tls_context_error serial=%s",
            request_id,
            serial_value,
        )
        raise HTTPException(
            status_code=500,
            detail=f"No se pudo inicializar TLS para SATO Cloud. request_id={request_id}",
        ) from exc

    def _dispatch_request() -> None:
        try:
            with urllib.request.urlopen(
                request,
                timeout=DEFAULT_SATO_CLOUD_TIMEOUT_SECONDS,
                context=tls_context,
            ):
                logger.info(
                    "labels.cloud[%s] async_ok serial=%s url=%s",
                    request_id,
                    serial_value,
                    target_url,
                )
                return
        except Exception:
            # Flujo beta: no bloqueamos el request del cliente por fallos async.
            logger.exception(
                "labels.cloud[%s] async_error serial=%s url=%s",
                request_id,
                serial_value,
                target_url,
            )
            return

    if payload.fire_and_forget:
        threading.Thread(target=_dispatch_request, daemon=True).start()
        return {
            "ok": True,
            "accepted": True,
            "mode": "fire_and_forget",
            "upstream_url": target_url,
            "request_id": request_id,
        }

    try:
        with urllib.request.urlopen(
            request,
            timeout=DEFAULT_SATO_CLOUD_TIMEOUT_SECONDS,
            context=tls_context,
        ) as upstream:
            response_body = upstream.read().decode("utf-8", errors="replace")
            return {
                "ok": True,
                "status": upstream.status,
                "upstream_url": target_url,
                "upstream_body": response_body,
                "request_id": request_id,
            }
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        logger.warning(
            "labels.cloud[%s] upstream_http_error serial=%s code=%s detail=%s",
            request_id,
            serial_value,
            exc.code,
            detail[:500],
        )
        raise HTTPException(
            status_code=502,
            detail=(
                f"SATO Cloud respondió {exc.code}: {detail or 'sin detalle'} "
                f"(request_id={request_id})"
            ),
        ) from exc
    except urllib.error.URLError as exc:
        reason_text = str(exc.reason)
        logger.warning(
            "labels.cloud[%s] upstream_url_error serial=%s reason=%s",
            request_id,
            serial_value,
            reason_text,
        )
        if "timed out" in reason_text.lower():
            raise HTTPException(
                status_code=504,
                detail=f"SATO Cloud tardó demasiado en responder. request_id={request_id}",
            ) from exc
        raise HTTPException(
            status_code=502,
            detail=f"No se pudo conectar a SATO Cloud: {reason_text} (request_id={request_id})",
        ) from exc
