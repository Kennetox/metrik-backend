from datetime import datetime, timedelta, timezone
from html import escape
import hashlib
from io import BytesIO
import os
from pathlib import Path
import re
from typing import List, Optional
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy import or_
from sqlalchemy.orm import Session

import crud
import models
import schemas
from database import get_db
from dependencies import require_any_permission, require_permission
from services import email as email_service
from services import monthly_report_email
from services import pdf_utils

router = APIRouter(
    prefix="/reports",
    tags=["reports"],
)


def _normalize_favorite_ids(ids: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in ids:
        preset_id = raw.strip()
        if not preset_id or len(preset_id) > 80 or preset_id in seen:
            continue
        seen.add(preset_id)
        normalized.append(preset_id)
    return normalized


def _favorites_version(preset_ids: list[str]) -> str:
    payload = "\x1f".join(preset_ids)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


class ReportPdfExportRequest(BaseModel):
    title: Optional[str] = None
    document_html: str
    preset_id: Optional[str] = None


@router.get(
    "/favorites",
    response_model=schemas.ReportFavoritesResponse,
)
def get_report_favorites(
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(
        require_any_permission("reports.view", "sales_history.view", "pos.sales")
    ),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    rows = (
        db.query(models.ReportFavorite.preset_id)
        .filter(models.ReportFavorite.tenant_id == tenant_id)
        .filter(models.ReportFavorite.user_id == current_user.id)
        .order_by(models.ReportFavorite.created_at.asc(), models.ReportFavorite.id.asc())
        .all()
    )
    preset_ids = _normalize_favorite_ids([row.preset_id for row in rows if row.preset_id])
    return schemas.ReportFavoritesResponse(
        preset_ids=preset_ids,
        version=_favorites_version(preset_ids),
    )


@router.put(
    "/favorites",
    response_model=schemas.ReportFavoritesResponse,
)
def update_report_favorites(
    payload: schemas.ReportFavoritesUpdateRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(
        require_any_permission("reports.view", "sales_history.view", "pos.sales")
    ),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    current_rows = (
        db.query(models.ReportFavorite.preset_id)
        .filter(models.ReportFavorite.tenant_id == tenant_id)
        .filter(models.ReportFavorite.user_id == current_user.id)
        .order_by(models.ReportFavorite.created_at.asc(), models.ReportFavorite.id.asc())
        .all()
    )
    current_ids = _normalize_favorite_ids(
        [row.preset_id for row in current_rows if row.preset_id]
    )
    current_version = _favorites_version(current_ids)

    if payload.expected_version and payload.expected_version != current_version:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Los favoritos fueron actualizados desde otra sesión.",
                "current_version": current_version,
            },
        )

    normalized = _normalize_favorite_ids(payload.preset_ids)

    (
        db.query(models.ReportFavorite)
        .filter(models.ReportFavorite.tenant_id == tenant_id)
        .filter(models.ReportFavorite.user_id == current_user.id)
        .delete(synchronize_session=False)
    )
    for preset_id in normalized:
        db.add(
            models.ReportFavorite(
                tenant_id=tenant_id,
                user_id=current_user.id,
                preset_id=preset_id,
            )
        )
    db.commit()
    return schemas.ReportFavoritesResponse(
        preset_ids=normalized,
        version=_favorites_version(normalized),
    )


def _month_utc_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    bogota_tz = ZoneInfo("America/Bogota")
    start_bogota = datetime(year, month, 1, tzinfo=bogota_tz)
    if month == 12:
        end_bogota = datetime(year + 1, 1, 1, tzinfo=bogota_tz)
    else:
        end_bogota = datetime(year, month + 1, 1, tzinfo=bogota_tz)
    return (
        start_bogota.astimezone(timezone.utc).replace(tzinfo=None),
        end_bogota.astimezone(timezone.utc).replace(tzinfo=None),
    )


def _normalize_group_key(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        return ""
    normalized = raw.replace("\\", "/").replace(">", "/").replace(" - ", "/")
    normalized = "/".join(part.strip() for part in normalized.split("/") if part.strip())
    return normalized


def _normalize_sku_key(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        return ""
    return "".join(ch for ch in raw if ch.isalnum())


def _normalize_name_key(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        return ""
    return "".join(ch for ch in raw if ch.isalnum() or ch.isspace()).strip()


def _sku_equivalent(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    if left.isdigit() and right.isdigit():
        return left.lstrip("0") == right.lstrip("0")
    return False


@router.post("/email")
def send_report_email(
    payload: schemas.ReportEmailRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("reports.view")),
):
    if not payload.recipients:
        raise HTTPException(status_code=400, detail="Debe indicar al menos un destinatario")
    if not payload.document_html:
        raise HTTPException(status_code=400, detail="El HTML del reporte es requerido")

    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    settings = crud.get_pos_settings(db, tenant_id=tenant_id)

    if payload.message:
        html_body = f"<p>{escape(payload.message)}</p>"
    else:
        html_body = "<p>Adjuntamos su reporte generado desde Kensar.</p>"

    attachments = []
    if payload.attach_pdf:
        if not pdf_utils.can_render_html_pdf():
            raise HTTPException(
                status_code=503,
                detail=(
                    "El servidor no tiene habilitada la generacion de PDF HTML "
                    "(dependencias de WeasyPrint faltantes)."
                ),
            )
        pdf_bytes = pdf_utils.build_pdf_from_html(payload.subject or "Reporte Kensar", payload.document_html)
        attachments.append(
            (
                "reporte_kensar.pdf",
                pdf_bytes,
                "application/pdf",
            )
        )

    try:
        email_service.send_email(
            recipients=payload.recipients,
            subject=payload.subject or "Reporte Kensar",
            html_body=html_body,
            attachments=attachments,
            smtp_config=settings,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except email_service.EmailDeliveryError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    return {"status": "sent"}


@router.post("/export/pdf")
def export_report_pdf(
    payload: ReportPdfExportRequest,
    _: object = Depends(require_permission("reports.view")),
):
    if not payload.document_html:
        raise HTTPException(status_code=400, detail="El HTML del reporte es requerido")
    if not pdf_utils.can_render_html_pdf():
        raise HTTPException(
            status_code=503,
            detail=(
                "El servidor no tiene habilitada la generacion de PDF HTML "
                "(dependencias de WeasyPrint faltantes)."
            ),
        )

    pdf_bytes = pdf_utils.build_pdf_from_html(payload.title or "Reporte Kensar", payload.document_html)
    safe_id = re.sub(r"[^a-z0-9_-]+", "_", (payload.preset_id or "reporte").lower()).strip("_") or "reporte"
    filename = f"{safe_id}_{datetime.now().strftime('%Y-%m-%d')}.pdf"

    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/monthly-quick/send-now",
    response_model=schemas.MonthlyQuickReportSendResponse,
)
def send_monthly_quick_report_now(
    payload: schemas.MonthlyQuickReportSendRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("settings.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    try:
        result = monthly_report_email.send_monthly_quick_report(
            db,
            tenant_id=tenant_id,
            year=payload.year,
            month=payload.month,
            force=payload.force,
            trigger="manual",
        )
        return schemas.MonthlyQuickReportSendResponse(**result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except email_service.EmailDeliveryError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get(
    "/quick/insights",
    response_model=schemas.ReportsQuickInsightsResponse,
)
def get_quick_insights(
    year: int,
    month: int,
    source: str = Query(default="all"),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(
        require_any_permission("reports.view", "sales_history.view", "pos.sales")
    ),
):
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Mes inválido")
    if year < 2000 or year > 2200:
        raise HTTPException(status_code=400, detail="Año inválido")

    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    normalized_source = (source or "all").strip().lower()
    include_metrik = normalized_source in {"all", "metrik"}
    include_legacy = normalized_source in {"all", "aronium", "legacy"}
    if normalized_source not in {"all", "metrik", "aronium", "legacy"}:
        raise HTTPException(status_code=400, detail="Filtro source inválido")
    start_utc, end_utc = _month_utc_bounds(year, month)

    min_sale_at = None
    max_sale_at = None
    if include_metrik:
        min_sale_at, max_sale_at = (
            db.query(
                func.min(models.Sale.created_at),
                func.max(models.Sale.created_at),
            )
            .filter(models.Sale.tenant_id == tenant_id)
            .filter(or_(models.Sale.status.is_(None), models.Sale.status != "voided"))
            .first()
            or (None, None)
        )
    if include_legacy:
        legacy_min, legacy_max = (
            db.query(
                func.min(models.LegacySale.created_at),
                func.max(models.LegacySale.created_at),
            )
            .filter(models.LegacySale.tenant_id == tenant_id)
            .filter(models.LegacySale.status == "completed")
            .first()
            or (None, None)
        )
        candidates_min = [value for value in [min_sale_at, legacy_min] if value is not None]
        candidates_max = [value for value in [max_sale_at, legacy_max] if value is not None]
        min_sale_at = min(candidates_min) if candidates_min else None
        max_sale_at = max(candidates_max) if candidates_max else None
    bogota_tz = ZoneInfo("America/Bogota")
    now_year = datetime.now(bogota_tz).year
    min_year = (
        min_sale_at.replace(tzinfo=timezone.utc).astimezone(bogota_tz).year
        if min_sale_at
        else now_year
    )
    max_year = (
        max_sale_at.replace(tzinfo=timezone.utc).astimezone(bogota_tz).year
        if max_sale_at
        else now_year
    )

    top_products_acc: dict[str, dict[str, float]] = {}
    top_groups_acc: dict[str, dict[str, float]] = {}

    if include_metrik:
        top_products_rows = (
            db.query(
                models.SaleItem.product_name.label("name"),
                func.sum(models.SaleItem.quantity).label("units"),
                func.sum(models.SaleItem.total).label("total"),
            )
            .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)
            .filter(models.Sale.tenant_id == tenant_id)
            .filter(or_(models.Sale.status.is_(None), models.Sale.status != "voided"))
            .filter(models.Sale.created_at >= start_utc)
            .filter(models.Sale.created_at < end_utc)
            .group_by(models.SaleItem.product_name)
            .all()
        )
        for row in top_products_rows:
            key = str(row.name or "Producto")
            bucket = top_products_acc.setdefault(key, {"units": 0.0, "total": 0.0})
            bucket["units"] += float(row.units or 0.0)
            bucket["total"] += float(row.total or 0.0)

        group_name_expr = func.coalesce(models.Product.group_name, "Sin grupo")
        top_groups_rows = (
            db.query(
                group_name_expr.label("name"),
                func.sum(models.SaleItem.quantity).label("units"),
                func.sum(models.SaleItem.total).label("total"),
            )
            .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)
            .outerjoin(models.Product, models.Product.id == models.SaleItem.product_id)
            .filter(models.Sale.tenant_id == tenant_id)
            .filter(or_(models.Sale.status.is_(None), models.Sale.status != "voided"))
            .filter(models.Sale.created_at >= start_utc)
            .filter(models.Sale.created_at < end_utc)
            .group_by(group_name_expr)
            .all()
        )
        for row in top_groups_rows:
            key = str(row.name or "Sin grupo")
            bucket = top_groups_acc.setdefault(key, {"units": 0.0, "total": 0.0})
            bucket["units"] += float(row.units or 0.0)
            bucket["total"] += float(row.total or 0.0)

    if include_legacy:
        legacy_product_rows = (
            db.query(
                models.LegacySaleItem.product_name.label("name"),
                func.sum(models.LegacySaleItem.quantity).label("units"),
                func.sum(models.LegacySaleItem.total).label("total"),
            )
            .join(models.LegacySale, models.LegacySale.id == models.LegacySaleItem.legacy_sale_id)
            .filter(models.LegacySale.tenant_id == tenant_id)
            .filter(models.LegacySale.status == "completed")
            .filter(models.LegacySale.created_at >= start_utc)
            .filter(models.LegacySale.created_at < end_utc)
            .group_by(models.LegacySaleItem.product_name)
            .all()
        )
        for row in legacy_product_rows:
            key = str(row.name or "Producto")
            bucket = top_products_acc.setdefault(key, {"units": 0.0, "total": 0.0})
            bucket["units"] += float(row.units or 0.0)
            bucket["total"] += float(row.total or 0.0)

        legacy_group_expr = func.coalesce(models.LegacySaleItem.product_group, "Sin grupo")
        legacy_group_rows = (
            db.query(
                legacy_group_expr.label("name"),
                func.sum(models.LegacySaleItem.quantity).label("units"),
                func.sum(models.LegacySaleItem.total).label("total"),
            )
            .join(models.LegacySale, models.LegacySale.id == models.LegacySaleItem.legacy_sale_id)
            .filter(models.LegacySale.tenant_id == tenant_id)
            .filter(models.LegacySale.status == "completed")
            .filter(models.LegacySale.created_at >= start_utc)
            .filter(models.LegacySale.created_at < end_utc)
            .group_by(legacy_group_expr)
            .all()
        )
        for row in legacy_group_rows:
            key = str(row.name or "Sin grupo")
            bucket = top_groups_acc.setdefault(key, {"units": 0.0, "total": 0.0})
            bucket["units"] += float(row.units or 0.0)
            bucket["total"] += float(row.total or 0.0)

    top_products = [
        schemas.ReportsQuickTopRow(name=name, units=values["units"], total=values["total"])
        for name, values in sorted(
            top_products_acc.items(),
            key=lambda item: item[1]["total"],
            reverse=True,
        )[:5]
    ]
    top_groups = [
        schemas.ReportsQuickTopRow(name=name, units=values["units"], total=values["total"])
        for name, values in sorted(
            top_groups_acc.items(),
            key=lambda item: item[1]["total"],
            reverse=True,
        )[:5]
    ]

    return schemas.ReportsQuickInsightsResponse(
        year=year,
        month=month,
        min_year=min_year,
        max_year=max_year,
        top_products=top_products,
        top_groups=top_groups,
    )


@router.post(
    "/products/last-sales",
    response_model=schemas.ReportProductsLastSalesResponse,
)
def get_products_last_sales(
    payload: schemas.ReportProductsLastSalesRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(
        require_any_permission("reports.view", "sales_history.view", "pos.sales")
    ),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    sale_ids = [int(value) for value in payload.sale_ids if isinstance(value, int) and value > 0]
    product_ids = [
        int(value)
        for value in payload.product_ids
        if isinstance(value, int) and value > 0
    ]
    if not sale_ids or not product_ids:
        return schemas.ReportProductsLastSalesResponse(rows=[])

    rows = (
        db.query(
            models.SaleItem.product_id.label("product_id"),
            func.max(models.Sale.created_at).label("last_sale_at"),
        )
        .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)
        .filter(models.Sale.tenant_id == tenant_id)
        .filter(models.Sale.id.in_(sale_ids))
        .filter(models.SaleItem.product_id.isnot(None))
        .filter(models.SaleItem.product_id.in_(product_ids))
        .group_by(models.SaleItem.product_id)
        .all()
    )

    return schemas.ReportProductsLastSalesResponse(
        rows=[
            schemas.ReportProductLastSaleRow(
                product_id=int(row.product_id),
                last_sale_at=row.last_sale_at,
            )
            for row in rows
            if row.product_id is not None and row.last_sale_at is not None
        ]
    )


@router.post(
    "/products/by-target",
    response_model=schemas.ReportProductsByTargetResponse,
)
def get_products_by_target(
    payload: schemas.ReportProductsByTargetRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(
        require_any_permission("reports.view", "sales_history.view", "pos.sales")
    ),
):
    if payload.date_from > payload.date_to:
        raise HTTPException(status_code=400, detail="Rango de fechas inválido")
    if payload.mode == "product" and (payload.product_id is None or payload.product_id <= 0):
        raise HTTPException(status_code=400, detail="product_id es requerido para modo producto")
    if payload.mode == "group" and not (payload.group_path or payload.group_name):
        raise HTTPException(status_code=400, detail="group_path o group_name es requerido para modo grupo")

    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    normalized_source = (payload.source or "all").strip().lower()
    include_metrik = normalized_source in {"all", "metrik"}
    include_legacy = normalized_source in {"all", "aronium", "legacy"}
    if normalized_source not in {"all", "metrik", "aronium", "legacy"}:
        raise HTTPException(status_code=400, detail="Filtro source inválido")

    # Reutiliza fronteras UTC del día de Bogotá para evitar desfases en enero por timezone.
    bogota_tz = ZoneInfo("America/Bogota")
    start_bogota = datetime(payload.date_from.year, payload.date_from.month, payload.date_from.day, tzinfo=bogota_tz)
    end_day = datetime(payload.date_to.year, payload.date_to.month, payload.date_to.day, tzinfo=bogota_tz)
    end_bogota = end_day + timedelta(days=1)
    start_utc = start_bogota.astimezone(timezone.utc).replace(tzinfo=None)
    end_utc = end_bogota.astimezone(timezone.utc).replace(tzinfo=None)

    target_group_path = (payload.group_path or "").strip().lower()
    target_group_name = (payload.group_name or "").strip().lower()
    normalized_target_group_path = _normalize_group_key(target_group_path)
    normalized_target_group_name = _normalize_group_key(target_group_name)
    target_product_sku = _normalize_sku_key(payload.product_sku)
    target_product_name = _normalize_name_key(payload.product_name)

    rows: list[dict] = []
    documents: set[str] = set()
    units_total = 0.0
    total_value = 0.0

    if include_metrik:
        group_expr = func.coalesce(models.Product.group_name, "")
        query = (
            db.query(
                models.Sale.id.label("sale_id"),
                models.Sale.sale_number.label("sale_number"),
                models.Sale.document_number.label("document_number"),
                models.Sale.created_at.label("sale_at"),
                models.Sale.pos_name.label("pos_name"),
                models.SaleItem.product_id.label("product_id"),
                models.SaleItem.product_sku.label("sku"),
                models.SaleItem.product_name.label("product"),
                group_expr.label("group_name"),
                models.SaleItem.quantity.label("units"),
                models.SaleItem.unit_price.label("unit_value"),
                models.SaleItem.total.label("total_value"),
            )
            .join(models.SaleItem, models.SaleItem.sale_id == models.Sale.id)
            .outerjoin(models.Product, models.Product.id == models.SaleItem.product_id)
            .filter(models.Sale.tenant_id == tenant_id)
            .filter(or_(models.Sale.status.is_(None), models.Sale.status != "voided"))
            .filter(models.Sale.created_at >= start_utc)
            .filter(models.Sale.created_at < end_utc)
            .filter(models.SaleItem.quantity > 0)
        )
        if payload.mode == "product" and payload.product_id is not None:
            query = query.filter(models.SaleItem.product_id == payload.product_id)
        metrik_rows = query.all()
        for row in metrik_rows:
            group_name = (row.group_name or "").strip()
            if payload.mode == "group":
                normalized_group = _normalize_group_key(group_name)
                by_path = bool(normalized_target_group_path) and (
                    normalized_group == normalized_target_group_path
                    or normalized_group.startswith(f"{normalized_target_group_path}/")
                )
                by_name = bool(normalized_target_group_name) and (
                    normalized_group == normalized_target_group_name
                    or normalized_group.endswith(f"/{normalized_target_group_name}")
                )
                if not (by_path or by_name):
                    continue
            unit_value = float(row.unit_value or 0.0)
            units = float(row.units or 0.0)
            line_total = float(row.total_value or (unit_value * units))
            document = row.document_number or (f"#{int(row.sale_number):04d}" if row.sale_number else None)
            rows.append(
                {
                    "sku": (row.sku or "—").strip() or "—",
                    "product": (row.product or "Producto sin nombre").strip() or "Producto sin nombre",
                    "group": group_name or "Sin grupo",
                    "units": units,
                    "unit_value": unit_value,
                    "total_value": line_total,
                    "sale_at": row.sale_at,
                    "document": document or "—",
                    "pos_name": row.pos_name or "Sin POS",
                    "document_key": f"metrik:{row.sale_id}",
                    "product_id": int(row.product_id) if row.product_id is not None else None,
                }
            )
            documents.add(f"metrik:{row.sale_id}")
            units_total += units
            total_value += line_total

    if include_legacy:
        query = (
            db.query(
                models.LegacySale.id.label("sale_id"),
                models.LegacySale.sale_number.label("sale_number"),
                models.LegacySale.display_document_number.label("document_number"),
                models.LegacySale.created_at.label("sale_at"),
                models.LegacySale.pos_name.label("pos_name"),
                models.LegacySaleItem.product_id.label("product_id"),
                models.LegacySaleItem.product_sku.label("sku"),
                models.LegacySaleItem.product_name.label("product"),
                models.LegacySaleItem.product_group.label("group_name"),
                models.LegacySaleItem.quantity.label("units"),
                models.LegacySaleItem.unit_price.label("unit_value"),
                models.LegacySaleItem.total.label("total_value"),
            )
            .join(models.LegacySaleItem, models.LegacySaleItem.legacy_sale_id == models.LegacySale.id)
            .filter(models.LegacySale.tenant_id == tenant_id)
            .filter(models.LegacySale.status == "completed")
            .filter(models.LegacySale.created_at >= start_utc)
            .filter(models.LegacySale.created_at < end_utc)
            .filter(models.LegacySaleItem.quantity > 0)
        )
        legacy_rows = query.all()
        for row in legacy_rows:
            group_name = (row.group_name or "").strip()
            if payload.mode == "product":
                row_product_id = int(row.product_id) if row.product_id is not None else None
                if row_product_id != payload.product_id:
                    row_sku = _normalize_sku_key(row.sku)
                    row_product = _normalize_name_key(row.product)
                    sku_match = bool(target_product_sku) and _sku_equivalent(
                        row_sku, target_product_sku
                    )
                    name_match = bool(target_product_name) and (
                        row_product == target_product_name
                        or row_product.startswith(target_product_name)
                        or target_product_name.startswith(row_product)
                    )
                    if not (sku_match or name_match):
                        continue
            if payload.mode == "group":
                normalized_group = _normalize_group_key(group_name)
                by_path = bool(normalized_target_group_path) and (
                    normalized_group == normalized_target_group_path
                    or normalized_group.startswith(f"{normalized_target_group_path}/")
                )
                by_name = bool(normalized_target_group_name) and (
                    normalized_group == normalized_target_group_name
                    or normalized_group.endswith(f"/{normalized_target_group_name}")
                )
                if not (by_path or by_name):
                    continue
            unit_value = float(row.unit_value or 0.0)
            units = float(row.units or 0.0)
            line_total = float(row.total_value or (unit_value * units))
            document = row.document_number or (f"#{int(row.sale_number):04d}" if row.sale_number else None)
            rows.append(
                {
                    "sku": (row.sku or "—").strip() or "—",
                    "product": (row.product or "Producto sin nombre").strip() or "Producto sin nombre",
                    "group": group_name or "Sin grupo",
                    "units": units,
                    "unit_value": unit_value,
                    "total_value": line_total,
                    "sale_at": row.sale_at,
                    "document": document or "—",
                    "pos_name": row.pos_name or "Sin POS",
                    "document_key": f"legacy:{row.sale_id}",
                    "product_id": int(row.product_id) if row.product_id is not None else None,
                }
            )
            documents.add(f"legacy:{row.sale_id}")
            units_total += units
            total_value += line_total

    if payload.result_mode == "grouped":
        grouped: dict[str, dict] = {}
        for row in rows:
            product_id = row.get("product_id")
            key = f"id:{product_id}" if product_id else f"n:{row['sku']}|{row['product']}"
            if key not in grouped:
                grouped[key] = {
                    "sku": row["sku"],
                    "product": row["product"],
                    "group": row["group"],
                    "units": 0.0,
                    "total_value": 0.0,
                    "last_sale_at": None,
                }
            acc = grouped[key]
            acc["units"] += float(row["units"])
            acc["total_value"] += float(row["total_value"])
            if acc["last_sale_at"] is None or row["sale_at"] > acc["last_sale_at"]:
                acc["last_sale_at"] = row["sale_at"]
        out_rows = [
            schemas.ReportProductsByTargetRow(
                sku=entry["sku"],
                product=entry["product"],
                group=entry["group"] or "Sin grupo",
                units=float(entry["units"]),
                unit_value=float(entry["total_value"] / entry["units"]) if entry["units"] > 0 else 0.0,
                total_value=float(entry["total_value"]),
                last_sale_at=entry["last_sale_at"],
            )
            for entry in sorted(
                grouped.values(),
                key=lambda value: (
                    value["last_sale_at"] or datetime.min,
                    value["total_value"],
                ),
                reverse=True,
            )
        ]
    else:
        out_rows = [
            schemas.ReportProductsByTargetRow(
                sku=row["sku"],
                product=row["product"],
                group=row["group"] or "Sin grupo",
                units=float(row["units"]),
                unit_value=float(row["unit_value"]),
                total_value=float(row["total_value"]),
                sale_at=row["sale_at"],
                document=row["document"],
                pos_name=row["pos_name"],
            )
            for row in sorted(
                rows,
                key=lambda value: (
                    value["sale_at"] or datetime.min,
                    value["total_value"],
                ),
                reverse=True,
            )
        ]

    return schemas.ReportProductsByTargetResponse(
        rows_count=len(out_rows),
        units=units_total,
        total_value=total_value,
        documents=len(documents),
        rows=out_rows,
    )


def _parse_money(raw_value: str) -> Optional[float]:
    cleaned = raw_value.strip()
    if not cleaned:
        return None
    negative = cleaned.startswith("-")
    cleaned = cleaned.replace("$", "").replace(" ", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        numeric = float(cleaned)
    except ValueError:
        return None
    return -abs(numeric) if negative else numeric


def _parse_percent(raw_value: str) -> Optional[float]:
    cleaned = raw_value.strip().replace("%", "").replace(",", ".")
    if not cleaned:
        return None
    try:
        numeric = float(cleaned)
    except ValueError:
        return None
    return numeric / 100.0


def _parse_integer(raw_value: str) -> Optional[int]:
    cleaned = re.sub(r"[^\d-]", "", raw_value.strip())
    if not cleaned:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _parse_date(raw_value: str) -> Optional[datetime]:
    text = raw_value.strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%y %H:%M",
    ]
    for pattern in formats:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _infer_cell_value(column_name: str, raw_value: str):
    normalized_col = column_name.lower().strip()
    text_value = (raw_value or "").strip()

    is_money_col = any(
        token in normalized_col
        for token in (
            "precio",
            "valor",
            "ventas",
            "total",
            "monto",
            "saldo",
            "recargo",
            "pagado",
            "ticket promedio",
        )
    )
    is_percent_col = "%" in normalized_col or any(
        token in normalized_col for token in ("participacion", "porcentaje")
    )
    is_count_col = any(
        token in normalized_col for token in ("cantidad", "tickets", "unidades")
    )
    is_date_col = normalized_col.startswith("fecha") or "ultima venta" in normalized_col

    if is_money_col:
        parsed = _parse_money(text_value)
        if parsed is not None:
            return parsed, '"$"#,##0', "right"
    if is_percent_col or "%" in text_value:
        parsed = _parse_percent(text_value)
        if parsed is not None:
            return parsed, "0.0%", "right"
    if is_count_col:
        parsed = _parse_integer(text_value)
        if parsed is not None:
            return parsed, "#,##0", "right"
    if is_date_col:
        parsed_date = _parse_date(text_value)
        if parsed_date is not None:
            if parsed_date.hour == 0 and parsed_date.minute == 0 and parsed_date.second == 0:
                return parsed_date, "DD/MM/YYYY", "center"
            return parsed_date, "DD/MM/YYYY HH:MM", "center"

    return text_value, None, "left"


def _resolve_ticket_logo_path(settings: Optional[models.PosSettings]) -> Optional[Path]:
    if settings is None:
        return None
    logo_url = (settings.ticket_logo_url or settings.logo_url or "").strip()
    if not logo_url:
        return None

    parsed = urlparse(logo_url)
    logo_path = parsed.path or logo_url
    logo_path = logo_path.strip()
    if not logo_path:
        return None

    # Evitamos SVG: openpyxl no lo inserta de forma nativa.
    lower_path = logo_path.lower()
    if lower_path.endswith(".svg"):
        return None

    logo_dir = Path(os.getenv("POS_LOGO_UPLOAD_DIR", "uploads/pos-logos"))
    public_prefix = os.getenv("POS_LOGO_PUBLIC_PATH", "/uploads/pos-logos").rstrip("/")

    filename: Optional[str] = None
    if logo_path.startswith(public_prefix + "/"):
        filename = logo_path[len(public_prefix) + 1 :]
    elif "/uploads/pos-logos/" in logo_path:
        filename = logo_path.split("/uploads/pos-logos/", 1)[1]
    elif "/" not in logo_path:
        filename = logo_path

    if not filename:
        return None

    candidate = (logo_dir / filename).resolve()
    try:
        candidate.relative_to(logo_dir.resolve())
    except ValueError:
        return None
    if not candidate.exists() or not candidate.is_file():
        return None
    return candidate


@router.post("/export/xlsx")
def export_report_xlsx(
    payload: schemas.ReportExportRequest,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("reports.view")),
):
    if not payload.table.columns:
        raise HTTPException(status_code=400, detail="La tabla del reporte no tiene columnas.")

    workbook = Workbook()
    try:
        workbook.calculation_properties.fullCalcOnLoad = True
    except AttributeError:
        pass

    sheet = workbook.active
    sheet.title = "Reporte"

    table_columns = len(payload.table.columns)
    total_columns = max(table_columns, 6)
    border_color = "D5DBE7"
    brand_primary = "0F766E"
    brand_primary_soft = "CCFBF1"
    table_header_fill = "ECFEFF"
    zebra_fill = "F8FAFC"

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    row_idx = 1
    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_columns)
    title_cell = sheet.cell(row=row_idx, column=1, value=payload.title)
    title_cell.font = Font(name="Calibri", size=17, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = PatternFill(start_color=brand_primary, end_color=brand_primary, fill_type="solid")
    row_idx_height = 28
    sheet.row_dimensions[row_idx].height = row_idx_height
    row_idx += 1

    sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_columns)
    generated_cell = sheet.cell(
        row=row_idx,
        column=1,
        value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    generated_cell.font = Font(name="Calibri", size=10, color="475569")
    generated_cell.alignment = Alignment(horizontal="left", vertical="center")
    generated_cell.fill = PatternFill(
        start_color=brand_primary_soft, end_color=brand_primary_soft, fill_type="solid"
    )
    row_idx += 2

    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    settings = crud.get_pos_settings(db, tenant_id=tenant_id)
    logo_path = _resolve_ticket_logo_path(settings)
    if logo_path is not None:
        try:
            logo_image = XlsxImage(str(logo_path))
            max_logo_width = 180
            if logo_image.width and logo_image.width > max_logo_width:
                ratio = max_logo_width / float(logo_image.width)
                logo_image.width = int(logo_image.width * ratio)
                logo_image.height = int(logo_image.height * ratio)
            anchor_col = get_column_letter(max(1, total_columns - 1))
            logo_image.anchor = f"{anchor_col}1"
            sheet.add_image(logo_image)
        except Exception:
            # Si el formato no es compatible, el reporte igual se exporta.
            pass

    section_cell = sheet.cell(row=row_idx, column=1, value="Informacion general")
    section_cell.font = Font(size=12, bold=True, color="0F172A")
    row_idx += 1
    company_rows = [
        ("Empresa", payload.company.name or "N/A"),
        ("Direccion", payload.company.address or "N/A"),
        ("Email", payload.company.email or "N/A"),
        ("Telefono", payload.company.phone or "N/A"),
    ]
    for label, value in company_rows:
        sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True, color="334155")
        value_cell = sheet.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(color="0F172A")
        row_idx += 1

    row_idx += 2

    sheet.cell(row=row_idx, column=1, value="Filtros aplicados").font = Font(
        size=12, bold=True, color="0F172A"
    )
    row_idx += 1
    filter_rows = [
        ("Desde", payload.filters.from_date),
        ("Hasta", payload.filters.to_date),
        ("POS", payload.filters.pos_filter),
        ("Metodo", payload.filters.method_filter),
        ("Vendedor", payload.filters.seller_filter),
    ]
    for label, value in filter_rows:
        sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True, color="334155")
        sheet.cell(row=row_idx, column=2, value=value or "Todos")
        row_idx += 1

    row_idx += 2
    if payload.summary:
        sheet.cell(row=row_idx, column=1, value="Resumen ejecutivo").font = Font(
            size=12, bold=True, color="0F172A"
        )
        row_idx += 1
        summary_header_fill = PatternFill(
            start_color=brand_primary_soft, end_color=brand_primary_soft, fill_type="solid"
        )
        for item in payload.summary:
            label_cell = sheet.cell(row=row_idx, column=1, value=item.label)
            label_cell.font = Font(bold=True, color="134E4A")
            label_cell.fill = summary_header_fill
            label_cell.border = thin_border
            value_cell = sheet.cell(row=row_idx, column=2, value=item.value)
            value_cell.font = Font(bold=True, color="0F172A")
            value_cell.fill = summary_header_fill
            value_cell.border = thin_border
            row_idx += 1
        row_idx += 2

    table_header_row = row_idx
    header_fill = PatternFill(start_color=table_header_fill, end_color=table_header_fill, fill_type="solid")
    for col_idx, column_name in enumerate(payload.table.columns, start=1):
        cell = sheet.cell(row=table_header_row, column=col_idx, value=column_name)
        cell.font = Font(bold=True, color=brand_primary)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx += 1
    for row_number, raw_row in enumerate(payload.table.rows, start=1):
        normalized_row = list(raw_row[: len(payload.table.columns)])
        if len(normalized_row) < len(payload.table.columns):
            normalized_row.extend([""] * (len(payload.table.columns) - len(normalized_row)))
        for col_idx, raw_value in enumerate(normalized_row, start=1):
            column_name = payload.table.columns[col_idx - 1]
            value, number_format, horizontal_align = _infer_cell_value(column_name, raw_value)
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format
            cell.alignment = Alignment(
                horizontal=horizontal_align, vertical="center", wrap_text=horizontal_align == "left"
            )
            cell.border = thin_border
            if row_number % 2 == 0:
                cell.fill = PatternFill(start_color=zebra_fill, end_color=zebra_fill, fill_type="solid")
        row_idx += 1

    if not payload.table.rows and payload.table.empty_message:
        sheet.merge_cells(
            start_row=row_idx,
            start_column=1,
            end_row=row_idx + 1,
            end_column=table_columns,
        )
        empty_cell = sheet.cell(row=row_idx, column=1, value=payload.table.empty_message)
        empty_cell.font = Font(italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_idx += 2

    if payload.table.rows:
        last_data_row = table_header_row + len(payload.table.rows)
        sheet.auto_filter.ref = f"A{table_header_row}:{get_column_letter(table_columns)}{last_data_row}"
        sheet.freeze_panes = f"A{table_header_row + 1}"

    for col_idx in range(1, table_columns + 1):
        column_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in sheet[column_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(12, max_len + 2), 42)

    if table_columns >= 8:
        sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    meta_sheet = workbook.create_sheet("Metadata")
    meta_sheet.append(["Campo", "Valor"])
    meta_sheet["A1"].font = Font(bold=True, color="FFFFFF")
    meta_sheet["B1"].font = Font(bold=True, color="FFFFFF")
    meta_sheet["A1"].fill = PatternFill(start_color=brand_primary, end_color=brand_primary, fill_type="solid")
    meta_sheet["B1"].fill = PatternFill(start_color=brand_primary, end_color=brand_primary, fill_type="solid")
    metadata_rows = [
        ("Preset ID", payload.preset_id),
        ("Titulo", payload.title),
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Desde", payload.filters.from_date),
        ("Hasta", payload.filters.to_date),
        ("POS", payload.filters.pos_filter),
        ("Metodo", payload.filters.method_filter),
        ("Vendedor", payload.filters.seller_filter),
        ("Filas de tabla", str(len(payload.table.rows))),
        ("Columnas de tabla", str(table_columns)),
    ]
    for row in metadata_rows:
        meta_sheet.append(list(row))
    meta_sheet.column_dimensions["A"].width = 22
    meta_sheet.column_dimensions["B"].width = 50
    meta_sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    safe_id = re.sub(r"[^a-z0-9_-]+", "_", payload.preset_id.lower()).strip("_") or "reporte"
    filename = f"reporte_{safe_id}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
