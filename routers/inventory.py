from datetime import datetime
from html import escape as html_escape
from typing import Dict, List

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session
from sqlalchemy import Integer, and_, case, cast, func, or_, true
from fastapi.responses import StreamingResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import crud, models, schemas
from database import get_db
from dependencies import require_permission
from services import pdf_utils


router = APIRouter(
    prefix="/inventory",
    tags=["inventory"],
)


def _format_cop_whole(value: float) -> str:
    rounded = int(round(float(value or 0.0)))
    abs_str = f"{abs(rounded):,}".replace(",", ".")
    return f"-${abs_str}" if rounded < 0 else f"${abs_str}"


def _sale_context_by_id(
    db: Session,
    sale_ids: List[int],
    tenant_id: int | None,
) -> Dict[int, Dict[str, str | None]]:
    if not sale_ids:
        return {}

    sale_rows = (
        db.query(
            models.Sale.id,
            models.Sale.pos_name,
            models.Sale.vendor_name,
            models.Sale.station_id,
        )
        .filter(models.Sale.id.in_(sale_ids))
        .filter(models.Sale.tenant_id == tenant_id if tenant_id is not None else true())
        .all()
    )

    station_ids = [row.station_id for row in sale_rows if row.station_id]
    station_label_by_id: Dict[str, str] = {}
    if station_ids:
        station_rows = (
            db.query(models.PosStation.id, models.PosStation.label)
            .filter(models.PosStation.id.in_(station_ids))
            .filter(
                models.PosStation.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .all()
        )
        station_label_by_id = {row.id: row.label for row in station_rows}

    out: Dict[int, Dict[str, str | None]] = {}
    for row in sale_rows:
        pos_name = row.pos_name or station_label_by_id.get(row.station_id or "")
        out[row.id] = {
            "sale_pos_name": pos_name,
            "sale_seller_name": row.vendor_name,
        }
    return out


def _resolve_receiving_entry_source(origin_name: str | None) -> str:
    normalized = (origin_name or "").strip().lower()
    if "web" in normalized or "metrik" in normalized:
        return "manual"
    return "app"


def _resolve_stock_device_or_422(
    db: Session,
    tenant_id: int | None,
    stock_device_id: str | None,
) -> models.StockDevice | None:
    cleaned = (stock_device_id or "").strip()
    if not cleaned:
        return None
    device = crud.get_stock_device(db, cleaned, tenant_id=tenant_id)
    if not device:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "DEVICE_NOT_ALLOWED",
                "message": "El dispositivo de inventario no existe para esta empresa.",
            },
        )
    if not device.is_active:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "DEVICE_BLOCKED",
                "message": "El dispositivo de inventario está inactivo.",
            },
        )
    return device


def _normalize_group_scope(value: str | None) -> str:
    return (value or "").strip().lower()


def _group_scope_matches(group_name: str | None, scope_value: str | None) -> bool:
    group = _normalize_group_scope(group_name)
    scope = _normalize_group_scope(scope_value)
    if not group or not scope:
        return False
    return group == scope or group.startswith(f"{scope}/")


@router.get("/overview", response_model=schemas.InventoryOverview)
def get_inventory_overview(
    status_limit: int = Query(default=6, ge=1, le=20),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label("qty_on_hand"),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )

    product_rows = (
        db.query(
            models.Product,
            func.coalesce(stock_subquery.c.qty_on_hand, 0).label("qty_on_hand"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .all()
    )

    total_qty = 0.0
    low_stock_count = 0
    critical_count = 0
    reorder_count = 0
    status_rows: List[schemas.InventoryStatusRow] = []

    for product, qty_on_hand in product_rows:
        qty = float(qty_on_hand or 0.0)
        total_qty += qty
        if qty <= 0:
            critical_count += 1
            status = "critical"
        elif product.low_stock_alert and product.stock_min > 0 and qty <= product.stock_min:
            low_stock_count += 1
            status = "low"
        else:
            status = "ok"

        if product.low_stock_alert and product.reorder_point > 0 and qty <= product.reorder_point:
            reorder_count += 1

        status_rows.append(
            schemas.InventoryStatusRow(
                product_id=product.id,
                product_name=product.name,
                qty_on_hand=qty,
                status=status,
            )
        )

    def status_rank(item: schemas.InventoryStatusRow) -> int:
        if item.status == "critical":
            return 0
        if item.status == "low":
            return 1
        return 2

    status_rows_sorted = sorted(status_rows, key=status_rank)[:status_limit]

    movement_rows = (
        db.query(models.InventoryMovement, models.Product.name, models.Product.sku)
        .join(models.Product, models.Product.id == models.InventoryMovement.product_id)
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .order_by(models.InventoryMovement.created_at.desc())
        .limit(8)
        .all()
    )

    recent_movements: List[schemas.InventoryMovementRead] = []
    sale_ids: List[int] = [
        int(movement.reference_id)
        for movement, _, _ in movement_rows
        if movement.reference_type == "sale" and movement.reference_id is not None
    ]
    sale_context = _sale_context_by_id(db, sale_ids, tenant_id)

    for movement, product_name, product_sku in movement_rows:
        sale_meta = (
            sale_context.get(int(movement.reference_id))
            if movement.reference_type == "sale" and movement.reference_id is not None
            else None
        )
        recent_movements.append(
            schemas.InventoryMovementRead(
                id=movement.id,
                product_id=movement.product_id,
                product_name=product_name,
                sku=product_sku,
                qty_delta=float(movement.qty_delta or 0.0),
                reason=movement.reason,
                notes=movement.notes,
                reference_type=movement.reference_type,
                reference_id=movement.reference_id,
                created_at=movement.created_at,
                created_by_user_id=movement.created_by_user_id,
                sale_pos_name=(sale_meta or {}).get("sale_pos_name"),
                sale_seller_name=(sale_meta or {}).get("sale_seller_name"),
            )
        )

    summary = schemas.InventorySummary(
        total_qty=total_qty,
        low_stock_count=low_stock_count,
        critical_count=critical_count,
        anomaly_count=0,
        reorder_count=reorder_count,
    )

    return schemas.InventoryOverview(
        summary=summary,
        recent_movements=recent_movements,
        status_rows=status_rows_sorted,
    )


@router.get("/movements", response_model=List[schemas.InventoryMovementRead])
def list_inventory_movements(
    skip: int = 0,
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    movement_rows = (
        db.query(models.InventoryMovement, models.Product.name, models.Product.sku)
        .join(models.Product, models.Product.id == models.InventoryMovement.product_id)
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .order_by(models.InventoryMovement.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    results: List[schemas.InventoryMovementRead] = []
    sale_ids: List[int] = [
        int(movement.reference_id)
        for movement, _, _ in movement_rows
        if movement.reference_type == "sale" and movement.reference_id is not None
    ]
    sale_context = _sale_context_by_id(db, sale_ids, tenant_id)

    for movement, product_name, product_sku in movement_rows:
        sale_meta = (
            sale_context.get(int(movement.reference_id))
            if movement.reference_type == "sale" and movement.reference_id is not None
            else None
        )
        results.append(
            schemas.InventoryMovementRead(
                id=movement.id,
                product_id=movement.product_id,
                product_name=product_name,
                sku=product_sku,
                qty_delta=float(movement.qty_delta or 0.0),
                reason=movement.reason,
                notes=movement.notes,
                reference_type=movement.reference_type,
                reference_id=movement.reference_id,
                created_at=movement.created_at,
                created_by_user_id=movement.created_by_user_id,
                sale_pos_name=(sale_meta or {}).get("sale_pos_name"),
                sale_seller_name=(sale_meta or {}).get("sale_seller_name"),
            )
        )

    return results


@router.get("/latest-entries", response_model=List[schemas.InventoryLatestEntryRead])
def list_latest_inventory_entries(
    source: str | None = Query(default="all", pattern="^(all|app|manual)$"),
    limit: int = Query(default=12, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)

    entries: List[schemas.InventoryLatestEntryRead] = []

    include_manual = source in (None, "all", "manual")
    include_app = source in (None, "all", "app")

    if include_manual:
        manual_rows = (
            db.query(models.InventoryMovement, models.Product.name, models.Product.sku)
            .join(models.Product, models.Product.id == models.InventoryMovement.product_id)
            .filter(
                models.InventoryMovement.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .filter(
                models.Product.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .filter(models.InventoryMovement.qty_delta > 0)
            .filter(models.InventoryMovement.reference_type.in_(["invoice", "cash"]))
            .order_by(models.InventoryMovement.created_at.desc())
            .limit(limit * 3)
            .all()
        )

        for movement, product_name, product_sku in manual_rows:
            entries.append(
                schemas.InventoryLatestEntryRead(
                    id=f"manual-{movement.id}",
                    source="manual",
                    product_id=movement.product_id,
                    product_name=product_name,
                    sku=product_sku,
                    qty_delta=float(movement.qty_delta or 0.0),
                    reason=movement.reason,
                    reference_type=movement.reference_type,
                    reference_id=movement.reference_id,
                    created_at=movement.created_at,
                )
            )

    if include_app:
        app_rows = (
            db.query(
                models.ReceivingLotItem,
                models.ReceivingLot.id.label("lot_id"),
                models.ReceivingLot.lot_number.label("lot_number"),
                models.ReceivingLot.closed_at.label("closed_at"),
                models.ReceivingLot.origin_name.label("origin_name"),
            )
            .join(models.ReceivingLot, models.ReceivingLot.id == models.ReceivingLotItem.lot_id)
            .filter(models.ReceivingLot.status == "closed")
            .filter(
                models.ReceivingLot.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .filter(
                models.ReceivingLotItem.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .order_by(models.ReceivingLot.closed_at.desc(), models.ReceivingLotItem.id.desc())
            .limit(limit * 3)
            .all()
        )

        for item, lot_id, lot_number, closed_at, origin_name in app_rows:
            created_at = closed_at or item.created_at
            resolved_source = _resolve_receiving_entry_source(origin_name)
            entries.append(
                schemas.InventoryLatestEntryRead(
                    id=f"app-{lot_id}-{item.id}",
                    source=resolved_source,
                    product_id=item.product_id,
                    product_name=item.product_name_snapshot,
                    sku=item.sku_snapshot,
                    qty_delta=float(item.qty_received or 0.0),
                    reason="purchase",
                    reference_type="receiving_lot",
                    reference_id=lot_id,
                    lot_id=lot_id,
                    lot_number=lot_number,
                    created_at=created_at,
                )
            )

    entries.sort(key=lambda item: item.created_at, reverse=True)
    return entries[:limit]


def _build_recount_summary(
    db: Session,
    recount_id: int,
    tenant_id: int | None,
) -> schemas.InventoryRecountSummary:
    summary_row = (
        db.query(
            func.count(models.InventoryRecountLine.id).label("total_lines"),
            func.coalesce(
                func.sum(
                    case(
                        (models.InventoryRecountLine.counted_qty.isnot(None), 1),
                        else_=0,
                    )
                ),
                0,
            ).label("counted_lines"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            and_(
                                models.InventoryRecountLine.counted_qty.isnot(None),
                                models.InventoryRecountLine.counted_qty
                                != models.InventoryRecountLine.system_qty,
                            ),
                            1,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("difference_lines"),
            func.coalesce(func.sum(models.InventoryRecountLine.system_qty), 0).label(
                "total_system_qty"
            ),
            func.coalesce(
                func.sum(
                    case(
                        (
                            models.InventoryRecountLine.counted_qty.isnot(None),
                            models.InventoryRecountLine.counted_qty,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("total_counted_qty"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            models.InventoryRecountLine.counted_qty.isnot(None),
                            models.InventoryRecountLine.counted_qty
                            - models.InventoryRecountLine.system_qty,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("total_diff_qty"),
        )
        .filter(models.InventoryRecountLine.recount_id == recount_id)
        .filter(
            models.InventoryRecountLine.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )

    total_lines = int(getattr(summary_row, "total_lines", 0) or 0)
    counted_lines = int(getattr(summary_row, "counted_lines", 0) or 0)
    pending_lines = max(0, total_lines - counted_lines)
    return schemas.InventoryRecountSummary(
        total_lines=total_lines,
        counted_lines=counted_lines,
        pending_lines=pending_lines,
        difference_lines=int(getattr(summary_row, "difference_lines", 0) or 0),
        total_system_qty=float(getattr(summary_row, "total_system_qty", 0.0) or 0.0),
        total_counted_qty=float(getattr(summary_row, "total_counted_qty", 0.0) or 0.0),
        total_diff_qty=float(getattr(summary_row, "total_diff_qty", 0.0) or 0.0),
    )


def _build_recount_read(
    db: Session,
    recount: models.InventoryRecount,
    tenant_id: int | None,
) -> schemas.InventoryRecountRead:
    summary = _build_recount_summary(db, recount.id, tenant_id)
    return schemas.InventoryRecountRead(
        id=recount.id,
        code=recount.code or f"RCN-{recount.id:06d}",
        status=recount.status,
        source=(recount.source or "web"),
        stock_device_id=recount.stock_device_id,
        stock_device_name=recount.stock_device_name,
        scope_type=recount.scope_type,
        scope_value=recount.scope_value,
        count_mode=recount.count_mode,
        title=recount.title,
        notes=recount.notes,
        created_by_user_id=recount.created_by_user_id,
        created_by_user_name=recount.created_by_user_name,
        closed_by_user_id=recount.closed_by_user_id,
        closed_by_user_name=recount.closed_by_user_name,
        applied_by_user_id=recount.applied_by_user_id,
        applied_by_user_name=recount.applied_by_user_name,
        created_at=recount.created_at,
        started_at=recount.started_at,
        closed_at=recount.closed_at,
        applied_at=recount.applied_at,
        cancelled_at=recount.cancelled_at,
        summary=summary,
    )


def _get_recount_or_404(
    db: Session,
    recount_id: int,
    tenant_id: int | None,
) -> models.InventoryRecount:
    recount = (
        db.query(models.InventoryRecount)
        .filter(models.InventoryRecount.id == recount_id)
        .filter(
            models.InventoryRecount.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    if not recount:
        raise HTTPException(status_code=404, detail="Recuento no encontrado")
    return recount


@router.get("/recounts", response_model=schemas.InventoryRecountPage)
def list_inventory_recounts(
    status_filter: schemas.InventoryRecountStatus | None = Query(
        default=None,
        alias="status",
    ),
    source_filter: schemas.InventoryRecountSource | None = Query(
        default=None,
        alias="source",
    ),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    base_query = db.query(models.InventoryRecount).filter(
        models.InventoryRecount.tenant_id == tenant_id
        if tenant_id is not None
        else true()
    )
    if status_filter:
        base_query = base_query.filter(models.InventoryRecount.status == status_filter)
    if source_filter:
        base_query = base_query.filter(models.InventoryRecount.source == source_filter)
    effective_date = func.coalesce(
        models.InventoryRecount.applied_at,
        models.InventoryRecount.closed_at,
        models.InventoryRecount.created_at,
    )
    if date_from:
        base_query = base_query.filter(effective_date >= date_from)
    if date_to:
        base_query = base_query.filter(effective_date < date_to)

    total = int(base_query.with_entities(func.count(models.InventoryRecount.id)).scalar() or 0)
    rows = (
        base_query.order_by(effective_date.desc(), models.InventoryRecount.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    items = [_build_recount_read(db, recount, tenant_id) for recount in rows]
    return schemas.InventoryRecountPage(items=items, total=total, skip=skip, limit=limit)


@router.post(
    "/recounts",
    response_model=schemas.InventoryRecountRead,
    status_code=status.HTTP_201_CREATED,
)
def create_inventory_recount(
    payload: schemas.InventoryRecountCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    pending_open_count = (
        db.query(func.count(models.InventoryRecount.id))
        .filter(
            models.InventoryRecount.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .filter(models.InventoryRecount.status.in_(["draft", "counting", "closed"]))
        .scalar()
        or 0
    )
    if pending_open_count >= 2:
        raise HTTPException(
            status_code=400,
            detail=(
                "Límite alcanzado: máximo 2 recuentos en curso/pedientes por empresa. "
                "Aplica o cancela uno antes de crear otro."
            ),
        )
    scope_value = payload.scope_value.strip() if payload.scope_value else None
    if payload.scope_type == "group" and not scope_value:
        raise HTTPException(
            status_code=400,
            detail="Debes indicar la categoría/grupo para recuento por grupo.",
        )
    stock_device = _resolve_stock_device_or_422(db, tenant_id, payload.stock_device_id)

    recount = models.InventoryRecount(
        tenant_id=tenant_id,
        status="counting",
        source=payload.source,
        stock_device_id=stock_device.id if stock_device else None,
        stock_device_name=stock_device.name if stock_device else None,
        scope_type=payload.scope_type,
        scope_value=scope_value,
        count_mode=payload.count_mode,
        title=(payload.title or "").strip() or None,
        notes=(payload.notes or "").strip() or None,
        created_by_user_id=current_user.id,
        started_at=datetime.utcnow(),
    )
    db.add(recount)
    db.flush()
    recount.code = f"RCN-{recount.id:06d}"

    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label("qty_on_hand"),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )

    products_query = (
        db.query(models.Product, func.coalesce(stock_subquery.c.qty_on_hand, 0).label("qty_on_hand"))
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.active.is_(True))
        .filter(models.Product.service.is_(False))
        .filter(models.Product.tenant_id == tenant_id if tenant_id is not None else true())
    )
    if payload.scope_type == "group" and scope_value:
        scope_norm = _normalize_group_scope(scope_value)
        products_query = products_query.filter(models.Product.group_name.isnot(None)).filter(
            or_(
                func.lower(models.Product.group_name) == scope_norm,
                func.lower(models.Product.group_name).like(f"{scope_norm}/%"),
            )
        )

    if payload.scope_type != "free":
        products_rows = products_query.order_by(models.Product.name.asc()).all()
        for product, qty_on_hand in products_rows:
            db.add(
                models.InventoryRecountLine(
                    tenant_id=tenant_id,
                    recount_id=recount.id,
                    product_id=product.id,
                    product_name_snapshot=product.name,
                    sku_snapshot=product.sku,
                    barcode_snapshot=product.barcode,
                    group_name_snapshot=product.group_name,
                    system_qty=float(qty_on_hand or 0.0),
                )
            )
    if stock_device:
        stock_device.last_seen_at = datetime.utcnow()

    db.commit()
    db.refresh(recount)
    return _build_recount_read(db, recount, tenant_id)


@router.get("/recounts/{recount_id}", response_model=schemas.InventoryRecountDetail)
def get_inventory_recount_detail(
    recount_id: int,
    q: str | None = Query(default=None, min_length=1),
    counted_only: bool = Query(default=False),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)

    line_query = (
        db.query(models.InventoryRecountLine)
        .filter(models.InventoryRecountLine.recount_id == recount.id)
        .filter(
            models.InventoryRecountLine.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    if q:
        pattern = f"%{q.strip()}%"
        line_query = line_query.filter(
            or_(
                models.InventoryRecountLine.product_name_snapshot.ilike(pattern),
                models.InventoryRecountLine.sku_snapshot.ilike(pattern),
                models.InventoryRecountLine.barcode_snapshot.ilike(pattern),
            )
        )
    if counted_only:
        line_query = line_query.filter(models.InventoryRecountLine.counted_qty.isnot(None))

    if counted_only:
        line_query = line_query.order_by(
            models.InventoryRecountLine.counted_at.desc(),
            models.InventoryRecountLine.updated_at.desc(),
        )
    else:
        line_query = line_query.order_by(models.InventoryRecountLine.product_name_snapshot.asc())
    line_rows = line_query.offset(skip).limit(limit).all()
    lines: List[schemas.InventoryRecountLineRead] = []
    for line in line_rows:
        diff_qty = (
            float(line.counted_qty - line.system_qty)
            if line.counted_qty is not None
            else None
        )
        lines.append(
            schemas.InventoryRecountLineRead(
                id=line.id,
                product_id=line.product_id,
                product_name=line.product_name_snapshot,
                sku=line.sku_snapshot,
                barcode=line.barcode_snapshot,
                group_name=line.group_name_snapshot,
                system_qty=float(line.system_qty or 0.0),
                counted_qty=float(line.counted_qty) if line.counted_qty is not None else None,
                diff_qty=diff_qty,
                notes=line.notes,
                counted_by_user_id=line.counted_by_user_id,
                counted_at=line.counted_at,
            )
        )

    return schemas.InventoryRecountDetail(
        recount=_build_recount_read(db, recount, tenant_id),
        lines=lines,
    )


@router.post(
    "/recounts/{recount_id}/lines",
    response_model=schemas.InventoryRecountLineRead,
)
def upsert_inventory_recount_line(
    recount_id: int,
    payload: schemas.InventoryRecountLineUpsert,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)
    if recount.status in ("applied", "cancelled"):
        raise HTTPException(status_code=400, detail="Este recuento no acepta cambios.")
    if payload.counted_qty < 0:
        raise HTTPException(status_code=400, detail="La cantidad contada no puede ser negativa.")

    line = (
        db.query(models.InventoryRecountLine)
        .filter(models.InventoryRecountLine.recount_id == recount.id)
        .filter(models.InventoryRecountLine.product_id == payload.product_id)
        .filter(
            models.InventoryRecountLine.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    if not line:
        product = (
            db.query(models.Product)
            .filter(models.Product.id == payload.product_id)
            .filter(models.Product.tenant_id == tenant_id if tenant_id is not None else true())
            .first()
        )
        if not product:
            raise HTTPException(status_code=404, detail="Producto no encontrado.")
        if not product.active or product.service:
            raise HTTPException(status_code=400, detail="Este producto no está disponible para recuento.")
        if recount.scope_type == "group" and not _group_scope_matches(product.group_name, recount.scope_value):
            raise HTTPException(
                status_code=400,
                detail="Este producto no pertenece a la categoría seleccionada para este recuento.",
            )

        system_qty = (
            db.query(func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0))
            .filter(models.InventoryMovement.product_id == product.id)
            .filter(models.InventoryMovement.tenant_id == tenant_id if tenant_id is not None else true())
            .scalar()
            or 0
        )
        line = models.InventoryRecountLine(
            tenant_id=tenant_id,
            recount_id=recount.id,
            product_id=product.id,
            product_name_snapshot=product.name,
            sku_snapshot=product.sku,
            barcode_snapshot=product.barcode,
            group_name_snapshot=product.group_name,
            system_qty=float(system_qty or 0.0),
        )
        db.add(line)
        db.flush()

    line.counted_qty = float(payload.counted_qty)
    line.notes = (payload.notes or "").strip() or None
    line.counted_by_user_id = current_user.id
    line.counted_at = datetime.utcnow()
    if recount.status == "draft":
        recount.status = "counting"
        recount.started_at = recount.started_at or datetime.utcnow()

    db.commit()
    db.refresh(line)

    diff_qty = (
        float(line.counted_qty - line.system_qty)
        if line.counted_qty is not None
        else None
    )
    return schemas.InventoryRecountLineRead(
        id=line.id,
        product_id=line.product_id,
        product_name=line.product_name_snapshot,
        sku=line.sku_snapshot,
        barcode=line.barcode_snapshot,
        group_name=line.group_name_snapshot,
        system_qty=float(line.system_qty or 0.0),
        counted_qty=float(line.counted_qty) if line.counted_qty is not None else None,
        diff_qty=diff_qty,
        notes=line.notes,
        counted_by_user_id=line.counted_by_user_id,
        counted_at=line.counted_at,
    )


@router.delete(
    "/recounts/{recount_id}/lines/{product_id}",
    response_model=schemas.InventoryRecountLineRead,
)
def clear_inventory_recount_line(
    recount_id: int,
    product_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)
    if recount.status in ("applied", "cancelled"):
        raise HTTPException(status_code=400, detail="Este recuento no acepta cambios.")

    line = (
        db.query(models.InventoryRecountLine)
        .filter(models.InventoryRecountLine.recount_id == recount.id)
        .filter(models.InventoryRecountLine.product_id == product_id)
        .filter(
            models.InventoryRecountLine.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    if not line:
        raise HTTPException(status_code=404, detail="Producto no incluido en este recuento.")

    line.counted_qty = None
    line.notes = None
    line.counted_by_user_id = None
    line.counted_at = None
    db.commit()
    db.refresh(line)

    return schemas.InventoryRecountLineRead(
        id=line.id,
        product_id=line.product_id,
        product_name=line.product_name_snapshot,
        sku=line.sku_snapshot,
        barcode=line.barcode_snapshot,
        group_name=line.group_name_snapshot,
        system_qty=float(line.system_qty or 0.0),
        counted_qty=None,
        diff_qty=None,
        notes=None,
        counted_by_user_id=None,
        counted_at=None,
    )


@router.post("/recounts/{recount_id}/close", response_model=schemas.InventoryRecountRead)
def close_inventory_recount(
    recount_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)
    if recount.status in ("applied", "cancelled"):
        raise HTTPException(status_code=400, detail="No se puede cerrar este recuento.")
    recount.status = "closed"
    recount.closed_by_user_id = current_user.id
    recount.closed_at = datetime.utcnow()
    db.commit()
    db.refresh(recount)
    return _build_recount_read(db, recount, tenant_id)


@router.post("/recounts/{recount_id}/cancel", response_model=schemas.InventoryRecountRead)
def cancel_inventory_recount(
    recount_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)
    if recount.status == "applied":
        raise HTTPException(status_code=400, detail="No se puede cancelar un recuento aplicado.")
    if recount.status == "cancelled":
        raise HTTPException(status_code=400, detail="Este recuento ya fue cancelado.")

    recount.status = "cancelled"
    recount.cancelled_at = datetime.utcnow()
    db.commit()
    db.refresh(recount)
    return _build_recount_read(db, recount, tenant_id)


@router.post("/recounts/{recount_id}/apply", response_model=schemas.InventoryRecountRead)
def apply_inventory_recount(
    recount_id: int,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    recount = _get_recount_or_404(db, recount_id, tenant_id)
    if recount.status == "applied":
        raise HTTPException(status_code=400, detail="Este recuento ya fue aplicado.")
    if recount.status == "cancelled":
        raise HTTPException(status_code=400, detail="No se puede aplicar un recuento cancelado.")
    if recount.status not in ("closed", "counting"):
        raise HTTPException(status_code=400, detail="Primero cierra el recuento para aplicarlo.")

    lines = (
        db.query(models.InventoryRecountLine)
        .filter(models.InventoryRecountLine.recount_id == recount.id)
        .filter(
            models.InventoryRecountLine.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .all()
    )

    adjustments = 0
    for line in lines:
        if line.counted_qty is None:
            continue
        diff_qty = float(line.counted_qty - line.system_qty)
        if abs(diff_qty) < 0.000001:
            continue
        movement = models.InventoryMovement(
            tenant_id=tenant_id,
            product_id=line.product_id,
            qty_delta=diff_qty,
            reason="count",
            notes=f"recuento:{recount.code or recount.id}" + (f" | {line.notes}" if line.notes else ""),
            reference_type="recount",
            reference_id=recount.id,
            created_by_user_id=current_user.id,
        )
        db.add(movement)
        adjustments += 1

    recount.status = "applied"
    recount.applied_by_user_id = current_user.id
    recount.applied_at = datetime.utcnow()
    if recount.closed_at is None:
        recount.closed_at = recount.applied_at
        recount.closed_by_user_id = current_user.id
    db.commit()
    db.refresh(recount)
    return _build_recount_read(db, recount, tenant_id)


def _apply_product_filters(
    query,
    qty_col,
    search: str | None,
    stock: str | None,
    group_name: str | None = None,
    status_filter: str | None = None,
):
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                models.Product.name.ilike(pattern),
                models.Product.sku.ilike(pattern),
                models.Product.barcode.ilike(pattern),
            )
        )

    if stock == "positive":
        query = query.filter(qty_col > 0)
    elif stock == "zero":
        query = query.filter(qty_col == 0)
    elif stock == "negative":
        query = query.filter(qty_col < 0)

    if group_name:
        query = query.filter(models.Product.group_name.ilike(f"%{group_name.strip()}%"))

    if status_filter == "negative":
        query = query.filter(qty_col < 0)
    elif status_filter == "critical":
        query = query.filter(qty_col <= 0)
    elif status_filter == "low":
        query = query.filter(
            qty_col > 0,
            models.Product.low_stock_alert.is_(True),
            models.Product.stock_min > 0,
            qty_col <= models.Product.stock_min,
        )
    elif status_filter == "ok":
        query = query.filter(
            qty_col > 0,
            or_(
                models.Product.low_stock_alert.is_(False),
                models.Product.low_stock_alert.is_(None),
                models.Product.stock_min <= 0,
                qty_col > models.Product.stock_min,
            ),
        )

    return query


def _apply_product_sort(query, qty_col, sort: str | None):
    sku_is_numeric = models.Product.sku.op("~")(r"^[0-9]+$")
    sku_numeric_value = cast(models.Product.sku, Integer)
    sku_numeric_rank = case((sku_is_numeric, 0), else_=1)
    cost_stock_value = qty_col * func.coalesce(models.Product.cost, 0)
    price_stock_value = qty_col * func.coalesce(models.Product.price, 0)

    if sort == "stock_asc":
        return query.order_by(qty_col.asc(), models.Product.name.asc())
    if sort == "stock_desc":
        return query.order_by(qty_col.desc(), models.Product.name.asc())
    if sort == "sku_asc":
        return query.order_by(
            sku_numeric_rank.asc(),
            sku_numeric_value.asc(),
            models.Product.sku.asc(),
            models.Product.name.asc(),
        )
    if sort == "sku_desc":
        return query.order_by(
            sku_numeric_rank.asc(),
            sku_numeric_value.desc(),
            models.Product.sku.desc(),
            models.Product.name.asc(),
        )
    if sort == "cost_stock_asc":
        return query.order_by(cost_stock_value.asc(), models.Product.name.asc())
    if sort == "cost_stock_desc":
        return query.order_by(cost_stock_value.desc(), models.Product.name.asc())
    if sort == "price_stock_asc":
        return query.order_by(price_stock_value.asc(), models.Product.name.asc())
    if sort == "price_stock_desc":
        return query.order_by(price_stock_value.desc(), models.Product.name.asc())
    return query.order_by(models.Product.name.asc())


@router.get("/products", response_model=schemas.InventoryProductPage)
def list_inventory_products(
    skip: int = 0,
    limit: int = Query(default=200, ge=1, le=1000),
    search: str | None = Query(default=None, min_length=1),
    group: str | None = Query(default=None, min_length=1),
    stock: str | None = Query(default=None, pattern="^(all|positive|zero|negative)$"),
    status_filter: str | None = Query(default=None, alias="status", pattern="^(all|ok|low|critical|negative)$"),
    sort: str | None = Query(
        default="name_asc",
        pattern="^(name_asc|stock_asc|stock_desc|sku_asc|sku_desc|cost_stock_asc|cost_stock_desc|price_stock_asc|price_stock_desc)$",
    ),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label(
                "qty_on_hand"
            ),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )
    latest_movement_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.max(models.InventoryMovement.created_at).label("last_movement_at"),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )

    qty_col = func.coalesce(stock_subquery.c.qty_on_hand, 0)

    product_rows = (
        db.query(
            models.Product,
            qty_col.label("qty_on_hand"),
            latest_movement_subquery.c.last_movement_at.label("last_movement_at"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .outerjoin(
            latest_movement_subquery,
            latest_movement_subquery.c.product_id == models.Product.id,
        )
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    status_value = None if status_filter in (None, "all") else status_filter
    product_rows = _apply_product_filters(product_rows, qty_col, search, stock, group, status_value)
    product_rows = _apply_product_sort(product_rows, qty_col, sort)
    product_rows = product_rows.offset(skip).limit(limit).all()

    count_query = (
        db.query(func.count(models.Product.id))
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    count_query = _apply_product_filters(count_query, qty_col, search, stock, group, status_value)
    total = int(count_query.scalar() or 0)

    totals_query = (
        db.query(
            func.coalesce(func.sum(qty_col * models.Product.cost), 0).label(
                "total_cost"
            ),
            func.coalesce(func.sum(qty_col * models.Product.price), 0).label(
                "total_price"
            ),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    totals_query = _apply_product_filters(totals_query, qty_col, search, stock, group, status_value)
    totals_row = totals_query.first()
    total_cost_value = float(getattr(totals_row, "total_cost", 0.0) or 0.0)
    total_price_value = float(getattr(totals_row, "total_price", 0.0) or 0.0)

    results: List[schemas.InventoryProductRow] = []
    for product, qty_on_hand, last_movement_at in product_rows:
        qty = float(qty_on_hand or 0.0)
        if qty <= 0:
            status = "critical"
        elif product.low_stock_alert and product.stock_min > 0 and qty <= product.stock_min:
            status = "low"
        else:
            status = "ok"

        results.append(
            schemas.InventoryProductRow(
                product_id=product.id,
                product_name=product.name,
                sku=product.sku,
                barcode=product.barcode,
                group_name=product.group_name,
                qty_on_hand=qty,
                status=status,
                cost=float(product.cost or 0.0),
                price=float(product.price or 0.0),
                last_movement_at=last_movement_at,
            )
        )

    return schemas.InventoryProductPage(
        items=results,
        total=int(total),
        skip=skip,
        limit=limit,
        total_cost_value=total_cost_value,
        total_price_value=total_price_value,
    )


@router.get("/products/export")
def export_inventory_products(
    search: str | None = Query(default=None, min_length=1),
    group: str | None = Query(default=None, min_length=1),
    stock: str | None = Query(default=None, pattern="^(all|positive|zero|negative)$"),
    status_filter: str | None = Query(default=None, alias="status", pattern="^(all|ok|low|critical|negative)$"),
    sort: str | None = Query(
        default="name_asc",
        pattern="^(name_asc|stock_asc|stock_desc|sku_asc|sku_desc|cost_stock_asc|cost_stock_desc|price_stock_asc|price_stock_desc)$",
    ),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label(
                "qty_on_hand"
            ),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )
    qty_col = func.coalesce(stock_subquery.c.qty_on_hand, 0)
    product_rows = (
        db.query(
            models.Product,
            qty_col.label("qty_on_hand"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    status_value = None if status_filter in (None, "all") else status_filter
    product_rows = _apply_product_filters(product_rows, qty_col, search, stock, group, status_value)
    product_rows = _apply_product_sort(product_rows, qty_col, sort).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "producto_id",
            "sku",
            "codigo_barras",
            "nombre",
            "categoria",
            "stock",
            "estado",
            "costo",
            "precio",
            "costo_en_stock",
            "precio_en_stock",
        ]
    )

    for product, qty_on_hand in product_rows:
        qty = float(qty_on_hand or 0.0)
        if qty <= 0:
            status = "critico"
        elif product.low_stock_alert and product.stock_min > 0 and qty <= product.stock_min:
            status = "bajo"
        else:
            status = "ok"
        writer.writerow(
            [
                product.id,
                product.sku or "",
                product.barcode or "",
                product.name,
                product.group_name or "",
                qty,
                status,
                int(round(float(product.cost or 0.0))),
                int(round(float(product.price or 0.0))),
                int(round(qty * float(product.cost or 0.0))),
                int(round(qty * float(product.price or 0.0))),
            ]
        )

    output.seek(0)
    response = StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
    )
    response.headers["Content-Disposition"] = "attachment; filename=inventario.csv"
    return response


@router.get("/products/export/xlsx")
def export_inventory_products_xlsx(
    search: str | None = Query(default=None, min_length=1),
    group: str | None = Query(default=None, min_length=1),
    stock: str | None = Query(default=None, pattern="^(all|positive|zero|negative)$"),
    status_filter: str | None = Query(default=None, alias="status", pattern="^(all|ok|low|critical|negative)$"),
    sort: str | None = Query(
        default="name_asc",
        pattern="^(name_asc|stock_asc|stock_desc|sku_asc|sku_desc|cost_stock_asc|cost_stock_desc|price_stock_asc|price_stock_desc)$",
    ),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label(
                "qty_on_hand"
            ),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )
    qty_col = func.coalesce(stock_subquery.c.qty_on_hand, 0)
    product_rows = (
        db.query(
            models.Product,
            qty_col.label("qty_on_hand"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    status_value = None if status_filter in (None, "all") else status_filter
    product_rows = _apply_product_filters(product_rows, qty_col, search, stock, group, status_value)
    product_rows = _apply_product_sort(product_rows, qty_col, sort).all()

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "producto_id",
        "sku",
        "codigo_barras",
        "nombre",
        "categoria",
        "stock",
        "estado",
        "costo",
        "precio",
        "costo_en_stock",
        "precio_en_stock",
    ]
    ws.append(headers)
    width_tracker: List[int] = [len(h) for h in headers]

    for product, qty_on_hand in product_rows:
        qty = float(qty_on_hand or 0.0)
        if qty <= 0:
            status = "critico"
        elif product.low_stock_alert and product.stock_min > 0 and qty <= product.stock_min:
            status = "bajo"
        else:
            status = "ok"

        unit_cost = float(product.cost or 0.0)
        unit_price = float(product.price or 0.0)
        row_values = [
            product.id,
            product.sku or "",
            product.barcode or "",
            product.name,
            product.group_name or "",
            qty,
            status,
            int(round(unit_cost)),
            int(round(unit_price)),
            int(round(qty * unit_cost)),
            int(round(qty * unit_price)),
        ]
        ws.append(row_values)

        display_values = [
            str(product.id),
            product.sku or "",
            product.barcode or "",
            product.name or "",
            product.group_name or "",
            str(int(round(qty))),
            status,
            _format_cop_whole(unit_cost),
            _format_cop_whole(unit_price),
            _format_cop_whole(qty * unit_cost),
            _format_cop_whole(qty * unit_price),
        ]
        for index, text in enumerate(display_values):
            width_tracker[index] = max(width_tracker[index], len(text))

    currency_number_format = '[$$-es-CO] #,##0'
    for row in range(2, ws.max_row + 1):
        ws[f"H{row}"].number_format = currency_number_format
        ws[f"I{row}"].number_format = currency_number_format
        ws[f"J{row}"].number_format = currency_number_format
        ws[f"K{row}"].number_format = currency_number_format

    # Auto-fit with sensible bounds so columns don't collapse or explode.
    for col_idx, tracked in enumerate(width_tracker, start=1):
        min_width = 10
        max_width = 44
        if col_idx in (1, 2, 6, 7):
            max_width = 14
        elif col_idx in (8, 9, 10, 11):
            min_width = 14
            max_width = 18
        elif col_idx in (3,):
            max_width = 24
        elif col_idx in (4, 5):
            max_width = 42
        width = min(max(tracked + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output)
    output.seek(0)
    headers = {
        "Content-Disposition": 'attachment; filename="inventario.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/products/export/pdf")
def export_inventory_products_pdf(
    search: str | None = Query(default=None, min_length=1),
    group: str | None = Query(default=None, min_length=1),
    stock: str | None = Query(default=None, pattern="^(all|positive|zero|negative)$"),
    status_filter: str | None = Query(default=None, alias="status", pattern="^(all|ok|low|critical|negative)$"),
    sort: str | None = Query(
        default="name_asc",
        pattern="^(name_asc|stock_asc|stock_desc|sku_asc|sku_desc|cost_stock_asc|cost_stock_desc|price_stock_asc|price_stock_desc)$",
    ),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    stock_subquery = (
        db.query(
            models.InventoryMovement.product_id.label("product_id"),
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label(
                "qty_on_hand"
            ),
        )
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .group_by(models.InventoryMovement.product_id)
        .subquery()
    )
    qty_col = func.coalesce(stock_subquery.c.qty_on_hand, 0)
    product_rows = (
        db.query(
            models.Product,
            qty_col.label("qty_on_hand"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.product_id == models.Product.id)
        .filter(models.Product.service.is_(False))
        .filter(models.Product.active.is_(True))
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    status_value = None if status_filter in (None, "all") else status_filter
    product_rows = _apply_product_filters(product_rows, qty_col, search, stock, group, status_value)
    product_rows = _apply_product_sort(product_rows, qty_col, sort).all()

    def _status_label(product: models.Product, qty: float) -> str:
        if qty < 0:
            return "Negativo"
        if qty <= 0:
            return "Crítico"
        if product.low_stock_alert and product.stock_min > 0 and qty <= product.stock_min:
            return "Bajo stock"
        return "Saludable"

    rows_html = ""
    for product, qty_on_hand in product_rows:
        qty = float(qty_on_hand or 0.0)
        unit_cost = float(product.cost or 0.0)
        unit_price = float(product.price or 0.0)
        cost_in_stock = qty * unit_cost
        price_in_stock = qty * unit_price
        negative_row_class = "neg" if qty < 0 else ""
        rows_html += f"""
        <tr class="{negative_row_class}">
          <td class="sku">{html_escape(product.sku or "-")}</td>
          <td class="prod"><span>{html_escape(product.name or "-")}</span></td>
          <td class="category">{html_escape(product.group_name or "-")}</td>
          <td class="num stock">{int(round(qty))}</td>
          <td class="num money unit-cost">{_format_cop_whole(unit_cost)}</td>
          <td class="num money unit-price">{_format_cop_whole(unit_price)}</td>
          <td class="num money total-cost">{_format_cop_whole(cost_in_stock)}</td>
          <td class="num money total-price">{_format_cop_whole(price_in_stock)}</td>
        </tr>
        """

    if not rows_html:
        rows_html = '<tr><td colspan="8" class="empty">Sin resultados para los filtros aplicados.</td></tr>'

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = f"""
    <!doctype html>
    <html>
      <head>
        <meta charset="utf-8" />
        <style>
          @page {{ size: A4 landscape; margin: 12mm; }}
          body {{ font-family: Arial, sans-serif; color: #0f172a; font-size: 10px; }}
          h1 {{ margin: 0 0 2px; font-size: 16px; }}
          .meta {{ margin: 0 0 8px; color: #475569; font-size: 9px; }}
          table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
          th, td {{ border: 1px solid #cbd5e1; padding: 4px 5px; word-wrap: break-word; }}
          th {{ background: #f1f5f9; text-align: left; font-size: 9px; }}
          thead {{ display: table-header-group; }}
          tfoot {{ display: table-row-group; }}
          tr {{ page-break-inside: avoid; break-inside: avoid; }}
          td, th {{ page-break-inside: avoid; break-inside: avoid; }}
          td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
          th.sku, td.sku {{ width: 4.5%; }}
          th.prod, td.prod {{ width: 18%; }}
          td.prod span {{ display: block; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 100%; }}
          th.category, td.category {{ width: 10%; }}
          th.stock, td.stock {{ width: 4.5%; }}
          th.money, td.money {{ width: 11.1%; }}
          th.money, td.money {{ padding-left: 4px; padding-right: 4px; }}
          td.money {{ font-size: 9px; }}
          tr.neg td.prod,
          tr.neg td.stock,
          tr.neg td.unit-cost,
          tr.neg td.unit-price,
          tr.neg td.total-cost,
          tr.neg td.total-price {{ color: #dc2626; font-weight: 700; }}
          .empty {{ text-align: center; color: #64748b; }}
        </style>
      </head>
      <body>
        <h1>Inventario exportado</h1>
        <p class="meta">Generado: {generated_at} · Registros: {len(product_rows)}</p>
        <table>
          <thead>
            <tr>
              <th class="sku">SKU</th>
              <th class="prod">Producto</th>
              <th class="category">Categoría</th>
              <th class="stock">Stock</th>
              <th class="money">Costo unit.</th>
              <th class="money">Precio unit.</th>
              <th class="money">Costo en stock</th>
              <th class="money">Precio en stock</th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
      </body>
    </html>
    """
    pdf_bytes = pdf_utils.build_pdf_from_html("Inventario", html)
    filename = f'inventario_{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/products/{product_id}/history", response_model=schemas.InventoryProductHistory)
def get_product_history(
    product_id: int,
    skip: int = 0,
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.view")),
):
    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    product = (
        db.query(models.Product)
        .filter(models.Product.id == product_id)
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    totals = (
        db.query(
            func.coalesce(func.sum(models.InventoryMovement.qty_delta), 0).label("net"),
            func.coalesce(
                func.sum(
                    case(
                        (models.InventoryMovement.qty_delta > 0, models.InventoryMovement.qty_delta),
                        else_=0,
                    )
                ),
                0,
            ).label("total_in"),
            func.coalesce(
                func.sum(
                    case(
                        (models.InventoryMovement.qty_delta < 0, -models.InventoryMovement.qty_delta),
                        else_=0,
                    )
                ),
                0,
            ).label("total_out"),
        )
        .filter(models.InventoryMovement.product_id == product_id)
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    net = float(getattr(totals, "net", 0.0) or 0.0)
    total_in = float(getattr(totals, "total_in", 0.0) or 0.0)
    total_out = float(getattr(totals, "total_out", 0.0) or 0.0)

    movement_query = (
        db.query(models.InventoryMovement)
        .filter(models.InventoryMovement.product_id == product_id)
        .filter(
            models.InventoryMovement.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
    )
    total_movements = int(
        movement_query.with_entities(func.count(models.InventoryMovement.id)).scalar()
        or 0
    )
    movement_rows = (
        movement_query.order_by(models.InventoryMovement.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    sale_ids: list[int] = []
    lot_ids: list[int] = []
    manual_doc_ids: list[int] = []
    recount_ids: list[int] = []
    for movement in movement_rows:
        if movement.reference_id is None:
            continue
        reference_type = (movement.reference_type or "").strip().lower()
        if reference_type == "sale":
            sale_ids.append(int(movement.reference_id))
        elif reference_type == "receiving_lot":
            lot_ids.append(int(movement.reference_id))
        elif reference_type in {"salida_manual", "venta_manual", "ajuste", "perdida_dano"}:
            manual_doc_ids.append(int(movement.reference_id))
        elif reference_type == "recount":
            recount_ids.append(int(movement.reference_id))

    sale_label_by_id: dict[int, str] = {}
    if sale_ids:
        sale_rows = (
            db.query(models.Sale.id, models.Sale.document_number, models.Sale.sale_number)
            .filter(models.Sale.id.in_(sale_ids))
            .filter(models.Sale.tenant_id == tenant_id if tenant_id is not None else true())
            .all()
        )
        for row in sale_rows:
            sale_label_by_id[row.id] = (
                row.document_number
                or (f"Ticket #{row.sale_number}" if row.sale_number is not None else f"Venta #{row.id}")
            )

    lot_label_by_id: dict[int, str] = {}
    if lot_ids:
        lot_rows = (
            db.query(models.ReceivingLot.id, models.ReceivingLot.lot_number)
            .filter(models.ReceivingLot.id.in_(lot_ids))
            .filter(
                models.ReceivingLot.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .all()
        )
        lot_label_by_id = {
            row.id: row.lot_number or f"Lote #{row.id}"
            for row in lot_rows
        }

    manual_label_by_id: dict[int, str] = {}
    if manual_doc_ids:
        manual_rows = (
            db.query(models.ManualMovementDocument.id, models.ManualMovementDocument.document_number)
            .filter(models.ManualMovementDocument.id.in_(manual_doc_ids))
            .filter(
                models.ManualMovementDocument.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .all()
        )
        manual_label_by_id = {
            row.id: row.document_number or f"Documento #{row.id}"
            for row in manual_rows
        }

    recount_label_by_id: dict[int, str] = {}
    if recount_ids:
        recount_rows = (
            db.query(models.InventoryRecount.id, models.InventoryRecount.code)
            .filter(models.InventoryRecount.id.in_(recount_ids))
            .filter(
                models.InventoryRecount.tenant_id == tenant_id
                if tenant_id is not None
                else true()
            )
            .all()
        )
        recount_label_by_id = {
            row.id: row.code or f"Recuento #{row.id}"
            for row in recount_rows
        }

    movements: List[schemas.InventoryProductMovement] = []
    for movement in movement_rows:
        reference_label: str | None = None
        if movement.reference_id is not None:
            ref_id = int(movement.reference_id)
            reference_type = (movement.reference_type or "").strip().lower()
            if reference_type == "sale":
                reference_label = sale_label_by_id.get(ref_id)
            elif reference_type == "receiving_lot":
                reference_label = lot_label_by_id.get(ref_id)
            elif reference_type in {"salida_manual", "venta_manual", "ajuste", "perdida_dano"}:
                reference_label = manual_label_by_id.get(ref_id)
            elif reference_type == "recount":
                reference_label = recount_label_by_id.get(ref_id)

        movements.append(
            schemas.InventoryProductMovement(
                id=movement.id,
                reason=movement.reason,
                qty_delta=float(movement.qty_delta or 0.0),
                notes=movement.notes,
                reference_type=movement.reference_type,
                reference_id=movement.reference_id,
                reference_label=reference_label,
                created_at=movement.created_at,
            )
        )

    return schemas.InventoryProductHistory(
        product_id=product.id,
        product_name=product.name,
        unit_cost=float(product.cost or 0.0),
        unit_price=float(product.price or 0.0),
        qty_on_hand=net,
        total_in=total_in,
        total_out=total_out,
        net=net,
        movements=movements,
        total_movements=total_movements,
        skip=skip,
        limit=limit,
    )


@router.post(
    "/movements",
    response_model=schemas.InventoryMovementRead,
    status_code=status.HTTP_201_CREATED,
)
def create_inventory_movement(
    payload: schemas.InventoryMovementCreate,
    db: Session = Depends(get_db),
    current_user: models.PosUser = Depends(require_permission("movements.manage")),
):
    if payload.qty_delta == 0:
        raise HTTPException(status_code=400, detail="La cantidad no puede ser 0")

    tenant_id = crud.resolve_user_tenant_id(db, current_user)
    product = (
        db.query(models.Product)
        .filter(models.Product.id == payload.product_id)
        .filter(
            models.Product.tenant_id == tenant_id
            if tenant_id is not None
            else true()
        )
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    if product.service:
        raise HTTPException(
            status_code=400,
            detail="No se pueden mover inventarios de productos tipo servicio",
        )

    movement = models.InventoryMovement(
        tenant_id=tenant_id,
        product_id=payload.product_id,
        qty_delta=payload.qty_delta,
        reason=payload.reason,
        notes=payload.notes,
        reference_type=payload.reference_type,
        reference_id=payload.reference_id,
        created_by_user_id=current_user.id,
    )
    db.add(movement)
    db.commit()
    db.refresh(movement)

    return schemas.InventoryMovementRead(
        id=movement.id,
        product_id=movement.product_id,
        product_name=product.name,
        sku=product.sku,
        qty_delta=float(movement.qty_delta or 0.0),
        reason=movement.reason,
        notes=movement.notes,
        reference_type=movement.reference_type,
        reference_id=movement.reference_id,
        created_at=movement.created_at,
        created_by_user_id=movement.created_by_user_id,
        sale_pos_name=None,
        sale_seller_name=None,
    )
